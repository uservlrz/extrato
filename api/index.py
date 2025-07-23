from http.server import BaseHTTPRequestHandler
import json
import pandas as pd
import io
import base64
import openpyxl
import re
import traceback

class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        
        response = {'status': 'OK', 'message': 'API de Extratos Bancários funcionando!'}
        self.wfile.write(json.dumps(response).encode())
    
    def do_POST(self):
        try:
            print("=== INICIANDO PROCESSAMENTO EXTRATOS ===")
            
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length)
            print(f"Dados recebidos: {len(post_data)} bytes")
            
            content_type = self.headers.get('Content-Type', '')
            if 'boundary=' not in content_type:
                raise Exception("Content-Type inválido - boundary não encontrado")
            
            boundary = content_type.split('boundary=')[1]
            print(f"Boundary: {boundary}")
            
            files, form_data = self.parse_multipart(post_data, boundary)
            print(f"Arquivos encontrados: {list(files.keys())}")
            print(f"Dados do formulário: {list(form_data.keys())}")
            
            csv_data = files.get('csv_file')
            excel_data = files.get('excel_file')
            
            if not csv_data or not excel_data:
                raise Exception("Arquivos CSV e Excel são necessários")
            
            print(f"CSV: {len(csv_data)} bytes")
            print(f"Excel: {len(excel_data)} bytes")
            print("Processando todas as transações (créditos + débitos)")
            
            # Processar Excel
            print("Processando Excel...")
            categorias = self.processar_excel(excel_data)
            print(f"Categorias encontradas: {len(categorias)}")
            
            # Processar CSV - SEMPRE incluir tudo
            print("Processando CSV...")
            df = self.processar_csv(csv_data, incluir_creditos=True)
            print(f"Linhas processadas: {len(df)}")
            print(f"Colunas: {list(df.columns)}")
            
            # Categorizar
            print("Categorizando transações...")
            df['Categoria'] = df['Descricao'].apply(lambda x: self.categorizar(x, categorias))
            
            # Separar créditos e débitos para análise
            df_creditos = df[df['Tipo'] == 'C'].copy()
            df_debitos = df[df['Tipo'] == 'D'].copy()
            
            print(f"Créditos: {len(df_creditos)} transações")
            print(f"Débitos: {len(df_debitos)} transações")
            
            # Agrupar resultados GERAIS (tudo junto)
            print("Agrupando resultados gerais...")
            resultados_gerais = df.groupby('Categoria').agg({
                'Valor': ['sum', 'count']
            }).reset_index()
            resultados_gerais.columns = ['categoria', 'total', 'quantidade']
            valor_total = df['Valor'].sum()
            
            if valor_total > 0:
                resultados_gerais['percentual'] = (resultados_gerais['total'] / valor_total) * 100
            else:
                resultados_gerais['percentual'] = 0
            resultados_gerais = resultados_gerais.sort_values('total', ascending=False)
            
            # Agrupar resultados CRÉDITOS
            print("Agrupando resultados de créditos...")
            if len(df_creditos) > 0:
                resultados_creditos = df_creditos.groupby('Categoria').agg({
                    'Valor': ['sum', 'count']
                }).reset_index()
                resultados_creditos.columns = ['categoria', 'total', 'quantidade']
                valor_total_creditos = df_creditos['Valor'].sum()
                
                if valor_total_creditos > 0:
                    resultados_creditos['percentual'] = (resultados_creditos['total'] / valor_total_creditos) * 100
                else:
                    resultados_creditos['percentual'] = 0
                resultados_creditos = resultados_creditos.sort_values('total', ascending=False)
            else:
                resultados_creditos = pd.DataFrame(columns=['categoria', 'total', 'quantidade', 'percentual'])
            
            # Agrupar resultados DÉBITOS
            print("Agrupando resultados de débitos...")
            if len(df_debitos) > 0:
                resultados_debitos = df_debitos.groupby('Categoria').agg({
                    'Valor': ['sum', 'count']
                }).reset_index()
                resultados_debitos.columns = ['categoria', 'total', 'quantidade']
                valor_total_debitos = df_debitos['Valor'].sum()
                
                if valor_total_debitos > 0:
                    resultados_debitos['percentual'] = (resultados_debitos['total'] / valor_total_debitos) * 100
                else:
                    resultados_debitos['percentual'] = 0
                resultados_debitos = resultados_debitos.sort_values('total', ascending=False)
            else:
                resultados_debitos = pd.DataFrame(columns=['categoria', 'total', 'quantidade', 'percentual'])
            
            # Função para preparar categorias detalhadas
            def preparar_categorias_detalhadas(resultados, dataframe):
                categorias_detalhadas = []
                for _, row in resultados.iterrows():
                    categoria = row['categoria']
                    itens_cat = dataframe[dataframe['Categoria'] == categoria]
                    
                    itens = []
                    for _, item in itens_cat.iterrows():
                        # Tratar valores None/NaN na data
                        data_valor = item['Data']
                        if pd.isna(data_valor):
                            data_formatada = None
                        else:
                            data_formatada = str(data_valor)
                        
                        itens.append({
                            'data': data_formatada,
                            'descricao': str(item['Descricao']),
                            'valor': float(item['Valor']),
                            'tipo': str(item['Tipo']),
                            'documento': str(item.get('Documento', ''))
                        })
                    
                    categorias_detalhadas.append({
                        'categoria': categoria,
                        'total': float(row['total']),
                        'quantidade': int(row['quantidade']),
                        'percentual': float(row['percentual']),
                        'itens': itens
                    })
                return categorias_detalhadas
            
            # Preparar respostas
            print("Preparando respostas...")
            categorias_gerais = preparar_categorias_detalhadas(resultados_gerais, df)
            categorias_creditos = preparar_categorias_detalhadas(resultados_creditos, df_creditos)
            categorias_debitos = preparar_categorias_detalhadas(resultados_debitos, df_debitos)
            
            # Gerar Excel
            print("Gerando Excel...")
            excel_b64 = self.gerar_excel_completo(categorias_gerais, categorias_creditos, categorias_debitos, df, df_creditos, df_debitos)
            
            resposta = {
                'success': True,
                'estatisticas': {
                    'total_transacoes': len(df),
                    'total_debitos': len(df_debitos),
                    'total_creditos': len(df_creditos),
                    'valor_total': float(valor_total),
                    'valor_total_creditos': float(df_creditos['Valor'].sum() if len(df_creditos) > 0 else 0),
                    'valor_total_debitos': float(df_debitos['Valor'].sum() if len(df_debitos) > 0 else 0)
                },
                'categorias_gerais': categorias_gerais,
                'categorias_creditos': categorias_creditos,
                'categorias_debitos': categorias_debitos,
                'excel_file': excel_b64
            }
            
            print("Enviando resposta...")
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(resposta).encode())
            print("=== PROCESSAMENTO CONCLUÍDO ===")
            
        except Exception as e:
            print(f"ERRO: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            error_response = {
                'success': False, 
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            self.wfile.write(json.dumps(error_response).encode())
    
    def parse_multipart(self, body, boundary):
        parts = body.split(f'--{boundary}'.encode())
        files = {}
        form_data = {}
        
        for part in parts:
            if b'Content-Disposition' not in part:
                continue
            
            header_end = part.find(b'\r\n\r\n')
            if header_end == -1:
                continue
            
            header = part[:header_end].decode('utf-8', errors='ignore')
            content = part[header_end + 4:].rstrip(b'\r\n-')
            
            if 'name="' in header:
                name_start = header.find('name="') + 6
                name_end = header.find('"', name_start)
                name = header[name_start:name_end]
                
                if 'filename="' in header:
                    files[name] = content
                else:
                    form_data[name] = content.decode('utf-8', errors='ignore')
        
        return files, form_data
    
    def processar_excel(self, excel_data):
        try:
            df = pd.read_excel(io.BytesIO(excel_data))
            
            if len(df.columns) < 2:
                raise Exception("Excel deve ter pelo menos 2 colunas")
            
            df.columns = ['Grupo', 'Palavra_Chave'] + list(df.columns[2:])
            
            categorias = {}
            categoria_atual = None
            
            for _, row in df.iterrows():
                if pd.notna(row['Grupo']) and str(row['Grupo']).strip():
                    categoria_atual = str(row['Grupo']).strip()
                
                if pd.notna(row['Palavra_Chave']) and categoria_atual:
                    palavra = str(row['Palavra_Chave']).strip()
                    categorias[palavra] = categoria_atual
            
            return categorias
        except Exception as e:
            raise Exception(f"Erro no Excel: {e}")
    
    def detectar_formato_csv(self, csv_string):
        """Detecta se é Banco do Brasil ou Bradesco"""
        try:
            print("=== DETECTANDO FORMATO ===")
            linhas = csv_string.split('\n')
            print(f"Analisando {len(linhas)} linhas...")
            
            bradesco_score = 0
            bb_score = 0
            
            # Analisar cada linha
            for i, linha in enumerate(linhas[:15]):
                linha_upper = linha.upper()
                
                # Indicadores do Bradesco
                if 'EXTRATO DE:' in linha_upper or 'AGÊNCIA:' in linha_upper or 'CONTA:' in linha_upper:
                    bradesco_score += 3
                
                if 'DATA;LANÇAMENTO;DCTO.' in linha_upper or 'DATA;LAN' in linha_upper:
                    bradesco_score += 3
                
                if linha.count('\r') > 5 and ';' in linha:
                    bradesco_score += 2
                
                if re.search(r'\d{2}/\d{2}/\d{4};.*?(PIX|CIELO|TRANSFERENCIA)', linha):
                    bradesco_score += 1
                
                # Indicadores do Banco do Brasil
                if '"DATA","DEPENDENCIA ORIGEM"' in linha_upper:
                    bb_score += 3
                
                if '"DATA"' in linha and '"HISTÓRICO"' in linha and '","' in linha:
                    bb_score += 3
                
                if linha.count('","') > 3 and linha.startswith('"'):
                    bb_score += 1
            
            print(f"Bradesco score: {bradesco_score}, BB score: {bb_score}")
            
            if bradesco_score >= bb_score and bradesco_score >= 2:
                print(f"FORMATO BRADESCO DETECTADO (score: {bradesco_score})")
                return 'bradesco'
            elif bb_score >= 2:
                print(f"FORMATO BANCO DO BRASIL DETECTADO (score: {bb_score})")
                return 'banco_brasil'
            
            # Heurísticas adicionais
            uso_pontovirgula = sum(linha.count(';') for linha in linhas[:10])
            uso_virgula = sum(linha.count(',') for linha in linhas[:10])
            
            if uso_pontovirgula > uso_virgula * 1.5:
                return 'bradesco'
            elif uso_virgula > uso_pontovirgula * 1.5:
                return 'banco_brasil'
            
            return 'desconhecido'
            
        except Exception as e:
            print(f"Erro na detecção de formato: {e}")
            return 'desconhecido'
    
    def processar_csv_bradesco(self, csv_string, incluir_creditos):
        """Processa CSV do Bradesco"""
        try:
            print("Processando CSV do Bradesco...")
            linhas = csv_string.split('\n')
            
            # Encontrar linha de dados
            linha_dados = None
            for i, linha in enumerate(linhas):
                if ('Data;Lançamento;Dcto.' in linha or 'Data;Lan' in linha) and len(linha) > 100:
                    linha_dados = linha
                    break
            
            if not linha_dados:
                inicio_dados = -1
                for i, linha in enumerate(linhas):
                    if 'Data;Lançamento;Dcto.' in linha or 'Data;Lan' in linha:
                        inicio_dados = i
                        break
                
                if inicio_dados >= 0:
                    linha_dados = ''.join(linhas[inicio_dados:])
                else:
                    raise Exception("Dados do Bradesco não encontrados")
            
            # Separar cabeçalho dos dados
            partes = linha_dados.split('\r')
            cabecalho = partes[0].strip()
            
            # Filtrar linhas válidas
            linhas_dados = []
            for parte in partes[1:]:
                linha_limpa = parte.strip()
                if (linha_limpa and 
                    not linha_limpa.startswith('Total;') and 
                    'SALDO ANTERIOR' not in linha_limpa and
                    ';' in linha_limpa and
                    linha_limpa.count(';') >= 4 and
                    re.match(r'^\d{2}/\d{2}/\d{4};', linha_limpa)):
                    linhas_dados.append(linha_limpa)
            
            if len(linhas_dados) == 0:
                raise Exception("Nenhuma linha válida encontrada")
            
            # Criar DataFrame
            csv_estruturado = cabecalho + '\n' + '\n'.join(linhas_dados)
            df = pd.read_csv(io.StringIO(csv_estruturado), delimiter=';')
            
            # Mapear colunas
            mapeamento = {}
            for col in df.columns:
                col_lower = col.lower().strip()
                if 'data' in col_lower:
                    mapeamento[col] = 'Data'
                elif 'lançamento' in col_lower or 'lancamento' in col_lower:
                    mapeamento[col] = 'Descricao'
                elif 'dcto' in col_lower:
                    mapeamento[col] = 'Documento'
                elif 'crédito' in col_lower or 'credito' in col_lower:
                    mapeamento[col] = 'Credito'
                elif 'débito' in col_lower or 'debito' in col_lower:
                    mapeamento[col] = 'Debito'
            
            df = df.rename(columns=mapeamento)
            
            # Garantir colunas
            if 'Credito' not in df.columns:
                df['Credito'] = 0.0
            if 'Debito' not in df.columns:
                df['Debito'] = 0.0
            if 'Documento' not in df.columns:
                df['Documento'] = ''
            
            # Processar valores
            def processar_valor(valor):
                if pd.isna(valor) or valor == '':
                    return 0.0
                valor_str = str(valor).strip().replace('.', '').replace(',', '.')
                try:
                    return abs(float(valor_str))
                except:
                    return 0.0
            
            df['Credito'] = df['Credito'].apply(processar_valor)
            df['Debito'] = df['Debito'].apply(processar_valor)
            
            # Determinar tipo e valor
            df['Tipo'] = df.apply(lambda row: 'C' if row['Credito'] > 0 else 'D', axis=1)
            df['Valor'] = df.apply(lambda row: row['Credito'] if row['Credito'] > 0 else row['Debito'], axis=1)
            
            # Limpar dados
            df = df.dropna(subset=['Descricao'])
            df = df[df['Valor'] > 0]
            
            # Processar datas
            try:
                df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y', errors='coerce')
            except:
                pass
            
            return df[['Data', 'Descricao', 'Valor', 'Tipo', 'Documento']].reset_index(drop=True)
            
        except Exception as e:
            raise Exception(f"Erro no processamento Bradesco: {e}")
    
    def processar_csv_banco_brasil(self, csv_string, incluir_creditos):
        """Processa CSV do Banco do Brasil"""
        # Limpar caracteres
        csv_string = csv_string.replace('Histórico', 'Historico')
        csv_string = csv_string.replace('Número', 'Numero')
        
        df = pd.read_csv(io.StringIO(csv_string))
        
        # Detectar formato
        if 'Descrição' in df.columns or 'Descricao' in df.columns:
            desc_col = 'Descrição' if 'Descrição' in df.columns else 'Descricao'
            df = df.dropna(subset=[desc_col, 'Valor'])
            df['Descricao'] = df[desc_col]
            df['Documento'] = df.get('Documento', '')
        elif 'Historico' in df.columns:
            df = df.dropna(subset=['Historico', 'Valor'])
            df = df[df['Historico'] != 'Saldo Anterior']
            df['Descricao'] = df['Historico']
            df['Documento'] = df.get('Numero do documento', '')
        else:
            raise Exception("Formato do BB não reconhecido")
        
        df['Tipo'] = df['Valor'].apply(lambda x: 'C' if x >= 0 else 'D')
        df['Valor'] = df['Valor'].abs()
        df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')
        df = df.dropna(subset=['Valor'])
        
        return df[['Data', 'Descricao', 'Valor', 'Tipo', 'Documento']].reset_index(drop=True)
    
    def processar_csv(self, csv_data, incluir_creditos):
        try:
            # Decodificar
            csv_string = None
            for encoding in ['utf-8', 'latin1', 'cp1252']:
                try:
                    csv_string = csv_data.decode(encoding)
                    break
                except:
                    continue
            
            if not csv_string:
                raise Exception("Não foi possível decodificar o CSV")
            
            # Detectar formato
            formato = self.detectar_formato_csv(csv_string)
            
            if formato == 'bradesco':
                return self.processar_csv_bradesco(csv_string, incluir_creditos)
            elif formato == 'banco_brasil':
                return self.processar_csv_banco_brasil(csv_string, incluir_creditos)
            else:
                raise Exception(f"Formato não reconhecido: {formato}")
                
        except Exception as e:
            raise Exception(f"Erro no CSV: {e}")
    
    def categorizar(self, descricao, categorias):
        if not descricao or pd.isna(descricao):
            return "Outros"
        
        desc_upper = str(descricao).upper()
        sorted_keys = sorted(categorias.keys(), key=len, reverse=True)
        
        for keyword in sorted_keys:
            if keyword.upper() in desc_upper:
                return categorias[keyword]
        
        return "Outros"
    
    def gerar_excel_completo(self, categorias_gerais, categorias_creditos, categorias_debitos, df_geral, df_creditos, df_debitos):
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # Aba Resumo
            ws_resumo = wb.create_sheet("Resumo Geral")
            ws_resumo.append(["ANÁLISE COMPLETA DE EXTRATO BANCÁRIO"])
            ws_resumo.append([f"Gerado em: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"])
            ws_resumo.append([])
            
            # Estatísticas
            total_transacoes = len(df_geral)
            total_debitos = len(df_debitos)
            total_creditos = len(df_creditos)
            valor_total = df_geral['Valor'].sum()
            valor_creditos = df_creditos['Valor'].sum() if len(df_creditos) > 0 else 0
            valor_debitos = df_debitos['Valor'].sum() if len(df_debitos) > 0 else 0
            
            ws_resumo.append(["ESTATÍSTICAS GERAIS"])
            ws_resumo.append(["Total de Transações", total_transacoes])
            ws_resumo.append(["Total de Créditos", total_creditos])
            ws_resumo.append(["Total de Débitos", total_debitos])
            ws_resumo.append(["Valor Total Geral", f"R$ {valor_total:,.2f}"])
            ws_resumo.append(["Valor Total Créditos", f"R$ {valor_creditos:,.2f}"])
            ws_resumo.append(["Valor Total Débitos", f"R$ {valor_debitos:,.2f}"])
            ws_resumo.append(["Saldo (Créditos - Débitos)", f"R$ {(valor_creditos - valor_debitos):,.2f}"])
            ws_resumo.append([])
            
            # Resumo por categoria
            ws_resumo.append(["RESUMO GERAL POR CATEGORIA"])
            ws_resumo.append(["Categoria", "Valor Total", "Quantidade", "Percentual"])
            
            for resultado in categorias_gerais:
                ws_resumo.append([
                    resultado['categoria'],
                    f"R$ {resultado['total']:,.2f}",
                    resultado['quantidade'],
                    f"{resultado['percentual']:.1f}%"
                ])
            
            # Abas adicionais se necessário
            if len(categorias_creditos) > 0:
                ws_creditos = wb.create_sheet("Resumo Créditos")
                ws_creditos.append(["ANÁLISE DE CRÉDITOS"])
                ws_creditos.append(["Categoria", "Valor", "Quantidade", "Percentual"])
                for resultado in categorias_creditos:
                    ws_creditos.append([
                        resultado['categoria'],
                        f"R$ {resultado['total']:,.2f}",
                        resultado['quantidade'],
                        f"{resultado['percentual']:.1f}%"
                    ])
            
            if len(categorias_debitos) > 0:
                ws_debitos = wb.create_sheet("Resumo Débitos")
                ws_debitos.append(["ANÁLISE DE DÉBITOS"])
                ws_debitos.append(["Categoria", "Valor", "Quantidade", "Percentual"])
                for resultado in categorias_debitos:
                    ws_debitos.append([
                        resultado['categoria'],
                        f"R$ {resultado['total']:,.2f}",
                        resultado['quantidade'],
                        f"{resultado['percentual']:.1f}%"
                    ])
            
            # Abas detalhadas por categoria
            for resultado in categorias_gerais:
                categoria = resultado['categoria']
                nome_aba = categoria.replace('/', '-').replace('\\', '-')[:31]
                ws_cat = wb.create_sheet(nome_aba)
                
                ws_cat.append([f"CATEGORIA: {categoria}"])
                ws_cat.append([f"Total: R$ {resultado['total']:,.2f}"])
                ws_cat.append([])
                ws_cat.append(["Data", "Descrição", "Valor", "Tipo"])
                
                for item in resultado['itens']:
                    data_formatada = str(item['data']) if item['data'] else 'Sem data'
                    tipo_formatado = "CRÉDITO" if item['tipo'] == 'C' else "DÉBITO"
                    ws_cat.append([
                        data_formatada,
                        item['descricao'],
                        f"R$ {item['valor']:,.2f}",
                        tipo_formatado
                    ])
            
            # Ajustar larguras
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Salvar
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return base64.b64encode(buffer.getvalue()).decode()
            
        except Exception as e:
            print(f"Erro ao gerar Excel: {e}")
            return None