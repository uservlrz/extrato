from http.server import BaseHTTPRequestHandler
import json
import pandas as pd
import io
import base64
import openpyxl
import traceback

class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        
        response = {'status': 'OK', 'message': 'API de Procedimentos Médicos funcionando!'}
        self.wfile.write(json.dumps(response).encode())
    
    def do_POST(self):
        try:
            print("=== INICIANDO PROCESSAMENTO PROCEDIMENTOS ===")
            
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length)
            
            content_type = self.headers.get('Content-Type', '')
            boundary = content_type.split('boundary=')[1]
            
            files, form_data = self.parse_multipart(post_data, boundary)
            print(f"Arquivos recebidos: {list(files.keys())}")
            
            # Obter arquivos
            procedures_data = files.get('procedures_file')
            categories_data = files.get('categories_file')
            
            if not procedures_data or not categories_data:
                raise Exception("Arquivos 'procedures_file' e 'categories_file' são obrigatórios")
            
            print(f"Procedimentos: {len(procedures_data)} bytes")
            print(f"Categorias: {len(categories_data)} bytes")
            
            # Processar categorias
            print("Processando categorias...")
            categorias = self.processar_categorias(categories_data)
            print(f"Categorias carregadas: {len(categorias)}")
            
            # Processar procedimentos
            print("Processando procedimentos...")
            df = self.processar_procedimentos_ipg(procedures_data)
            print(f"Procedimentos processados: {len(df)}")
            
            # Analisar dados
            print("Analisando dados...")
            analise = self.analisar_procedimentos(df, categorias)
            
            # Gerar Excel
            print("Gerando Excel...")
            excel_b64 = self.gerar_excel(analise, df)
            
            # Resposta
            resposta = {
                'success': True,
                'estatisticas': analise['estatisticas'],
                'categorias': analise['categorias'],
                'procedimentos': analise['procedimentos'],
                'unidades': analise['unidades'],
                'excel_file': excel_b64
            }
            
            print("Enviando resposta...")
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(resposta).encode())
            print("=== PROCESSAMENTO CONCLUÍDO ===")
            
        except Exception as e:
            print(f"ERRO: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            error_response = {
                'success': False,
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            self.wfile.write(json.dumps(error_response).encode())
    
    def parse_multipart(self, body, boundary):
        """Parse multipart form data"""
        parts = body.split(f'--{boundary}'.encode())
        files = {}
        form_data = {}
        
        for part in parts:
            if b'Content-Disposition' not in part:
                continue
            
            header_end = part.find(b'\r\n\r\n')
            if header_end == -1:
                continue
            
            header = part[:header_end].decode('utf-8', errors='ignore')
            content = part[header_end + 4:].rstrip(b'\r\n-')
            
            if 'name="' in header:
                name_start = header.find('name="') + 6
                name_end = header.find('"', name_start)
                name = header[name_start:name_end]
                
                if 'filename="' in header:
                    files[name] = content
                else:
                    form_data[name] = content.decode('utf-8', errors='ignore')
        
        return files, form_data
    
    def processar_categorias(self, data):
        """Processa arquivo Excel de categorias"""
        try:
            df = pd.read_excel(io.BytesIO(data))
            categorias = []
            
            for _, row in df.iterrows():
                if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip():
                    categorias.append(str(row.iloc[0]).strip())
            
            return categorias
            
        except Exception as e:
            raise Exception(f"Erro ao processar categorias: {e}")
    
    def processar_procedimentos_ipg(self, data):
        """Processa arquivo Excel de procedimentos IPG"""
        try:
            # Carregar arquivo
            df = pd.read_excel(io.BytesIO(data))
            print(f"Arquivo carregado: {df.shape}")
            
            # Estrutura IPG: Linha 2 = cabeçalhos, colunas 0, 5, 10
            # Encontrar linha de cabeçalhos
            header_row = 2  # Padrão IPG
            for i in range(min(5, len(df))):
                if pd.notna(df.iloc[i, 0]) and 'unidade' in str(df.iloc[i, 0]).lower():
                    header_row = i
                    print(f"Cabeçalhos encontrados na linha {i}")
                    break
            
            # Recarregar com cabeçalho correto
            df = pd.read_excel(io.BytesIO(data), header=header_row)
            print(f"Recarregado com header {header_row}: {df.shape}")
            print(f"Colunas: {list(df.columns)}")
            
            # Usar estrutura fixa do IPG: Unidade(0), Procedimento(5), Total(10)
            if df.shape[1] < 11:
                raise Exception(f"Arquivo deve ter pelo menos 11 colunas. Encontradas: {df.shape[1]}")
            
            # Extrair colunas pela posição
            df_proc = df.iloc[:, [0, 5, 10]].copy()
            df_proc.columns = ['Unidade', 'Procedimento', 'TotalItem']
            
            print(f"Dados extraídos: {len(df_proc)} linhas")
            print("Amostra dos dados:")
            print(df_proc.head().to_string())
            
            # Limpar dados
            # 1. Remover nulos
            df_proc = df_proc.dropna(subset=['Procedimento', 'TotalItem'])
            print(f"Após remover nulos: {len(df_proc)}")
            
            # 2. Remover procedimentos vazios
            df_proc = df_proc[df_proc['Procedimento'].astype(str).str.strip() != '']
            print(f"Após remover procedimentos vazios: {len(df_proc)}")
            
            # 3. Converter valores para numérico
            def converter_valor(valor):
                if pd.isna(valor):
                    return 0.0
                try:
                    if isinstance(valor, (int, float)):
                        return float(valor)
                    
                    valor_str = str(valor).strip()
                    # Remover símbolos monetários
                    valor_str = valor_str.replace('R, '').replace(', '').replace(' ', '')
                    # Trocar vírgula por ponto
                    if ',' in valor_str and '.' not in valor_str:
                        valor_str = valor_str.replace(',', '.')
                    
                    return float(valor_str)
                except:
                    return 0.0
            
            df_proc['TotalItem'] = df_proc['TotalItem'].apply(converter_valor)
            print(f"Valores convertidos - Min: {df_proc['TotalItem'].min()}, Max: {df_proc['TotalItem'].max()}")
            
            # 4. Filtrar valores > 0
            df_proc = df_proc[df_proc['TotalItem'] > 0]
            print(f"Após filtrar valores > 0: {len(df_proc)}")
            
            # 5. Limpar campos texto
            df_proc['Unidade'] = df_proc['Unidade'].fillna('Não informado').astype(str).str.strip()
            df_proc['Procedimento'] = df_proc['Procedimento'].astype(str).str.strip()
            
            if len(df_proc) == 0:
                raise Exception("Nenhum registro válido encontrado após limpeza")
            
            print(f"✅ Processamento concluído: {len(df_proc)} registros")
            print(f"Unidades únicas: {df_proc['Unidade'].nunique()}")
            print(f"Procedimentos únicos: {df_proc['Procedimento'].nunique()}")
            print(f"Valor total: R$ {df_proc['TotalItem'].sum():,.2f}")
            
            return df_proc
            
        except Exception as e:
            raise Exception(f"Erro ao processar procedimentos: {e}")
    
    def mapear_categoria(self, procedimento, categorias):
        """Mapeia procedimento para categoria"""
        if not procedimento:
            return "Outros"
        
        proc_upper = str(procedimento).upper()
        
        # Buscar correspondência exata
        for categoria in categorias:
            cat_upper = str(categoria).upper().strip()
            if cat_upper in proc_upper:
                return categoria
        
        # Verificações específicas
        for categoria in categorias:
            cat_upper = str(categoria).upper().strip()
            if cat_upper == 'VITAMINA' and ('VITAMINA' in proc_upper or 'B12' in proc_upper):
                return categoria
            if cat_upper == 'CONSULTAS' and 'CONSULTA' in proc_upper:
                return categoria
        
        return "Outros"
    
    def analisar_procedimentos(self, df, categorias):
        """Analisa procedimentos e gera relatórios"""
        try:
            # Categorizar procedimentos
            df['Categoria'] = df['Procedimento'].apply(lambda x: self.mapear_categoria(x, categorias))
            
            total_geral = df['TotalItem'].sum()
            
            # === ANÁLISE POR CATEGORIA ===
            cat_agrupadas = df.groupby('Categoria').agg({
                'TotalItem': ['sum', 'count']
            }).reset_index()
            cat_agrupadas.columns = ['Categoria', 'Total', 'Quantidade']
            cat_agrupadas['Percentual'] = (cat_agrupadas['Total'] / total_geral) * 100
            cat_agrupadas = cat_agrupadas.sort_values('Total', ascending=False)
            
            categorias_detalhadas = []
            for _, cat in cat_agrupadas.iterrows():
                # Procedimentos desta categoria
                procs_cat = df[df['Categoria'] == cat['Categoria']]
                proc_agrupados = procs_cat.groupby('Procedimento')['TotalItem'].agg(['sum', 'count']).reset_index()
                proc_agrupados.columns = ['Procedimento', 'Total', 'Quantidade']
                proc_agrupados = proc_agrupados.sort_values('Total', ascending=False)
                
                procedimentos_lista = []
                for _, proc in proc_agrupados.iterrows():
                    procedimentos_lista.append({
                        'procedimento': proc['Procedimento'],
                        'total': float(proc['Total']),
                        'quantidade': int(proc['Quantidade'])
                    })
                
                categorias_detalhadas.append({
                    'categoria': cat['Categoria'],
                    'total': float(cat['Total']),
                    'quantidade': int(cat['Quantidade']),
                    'percentual': float(cat['Percentual']),
                    'procedimentos': procedimentos_lista
                })
            
            # === ANÁLISE POR PROCEDIMENTO ===
            proc_agrupados = df.groupby('Procedimento').agg({
                'TotalItem': ['sum', 'count'],
                'Categoria': 'first'
            }).reset_index()
            proc_agrupados.columns = ['Procedimento', 'Total', 'Quantidade', 'Categoria']
            proc_agrupados = proc_agrupados.sort_values('Total', ascending=False)
            
            procedimentos_detalhados = []
            for _, proc in proc_agrupados.iterrows():
                # Unidades para este procedimento
                unidades_proc = df[df['Procedimento'] == proc['Procedimento']].groupby('Unidade')['TotalItem'].agg(['sum', 'count']).reset_index()
                unidades_proc.columns = ['Unidade', 'Total', 'Quantidade']
                
                unidades_dict = {}
                for _, unidade in unidades_proc.iterrows():
                    unidades_dict[unidade['Unidade']] = {
                        'total': float(unidade['Total']),
                        'quantidade': int(unidade['Quantidade'])
                    }
                
                procedimentos_detalhados.append({
                    'procedimento': proc['Procedimento'],
                    'categoria': proc['Categoria'],
                    'total': float(proc['Total']),
                    'quantidade': int(proc['Quantidade']),
                    'unidades': unidades_dict
                })
            
            # === ANÁLISE POR UNIDADE ===
            uni_agrupadas = df.groupby('Unidade').agg({
                'TotalItem': ['sum', 'count']
            }).reset_index()
            uni_agrupadas.columns = ['Unidade', 'Total', 'Quantidade']
            uni_agrupadas['Percentual'] = (uni_agrupadas['Total'] / total_geral) * 100
            uni_agrupadas = uni_agrupadas.sort_values('Total', ascending=False)
            
            unidades_detalhadas = []
            for _, unidade in uni_agrupadas.iterrows():
                # Categorias desta unidade
                cats_unidade = df[df['Unidade'] == unidade['Unidade']].groupby('Categoria')['TotalItem'].agg(['sum', 'count']).reset_index()
                cats_unidade.columns = ['Categoria', 'Total', 'Quantidade']
                cats_unidade = cats_unidade.sort_values('Total', ascending=False)
                
                categorias_lista = []
                for _, cat in cats_unidade.iterrows():
                    categorias_lista.append({
                        'categoria': cat['Categoria'],
                        'total': float(cat['Total']),
                        'quantidade': int(cat['Quantidade'])
                    })
                
                unidades_detalhadas.append({
                    'unidade': unidade['Unidade'],
                    'total': float(unidade['Total']),
                    'quantidade': int(unidade['Quantidade']),
                    'percentual': float(unidade['Percentual']),
                    'categorias': categorias_lista
                })
            
            # Estatísticas gerais
            estatisticas = {
                'total_procedimentos': len(df),
                'total_categorias': len(cat_agrupadas),
                'valor_total': float(total_geral),
                'total_unidades': len(uni_agrupadas)
            }
            
            return {
                'estatisticas': estatisticas,
                'categorias': categorias_detalhadas,
                'procedimentos': procedimentos_detalhados,
                'unidades': unidades_detalhadas
            }
            
        except Exception as e:
            raise Exception(f"Erro na análise: {e}")
    
    def gerar_excel(self, analise, df):
        """Gera Excel com relatório de procedimentos"""
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # === ABA RESUMO ===
            ws_resumo = wb.create_sheet("Resumo Geral")
            ws_resumo.append(["ANÁLISE DE PROCEDIMENTOS MÉDICOS IPG"])
            ws_resumo.append([f"Gerado em: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"])
            ws_resumo.append([])
            
            # Estatísticas
            stats = analise['estatisticas']
            ws_resumo.append(["ESTATÍSTICAS GERAIS"])
            ws_resumo.append(["Total de Procedimentos", stats['total_procedimentos']])
            ws_resumo.append(["Total de Categorias", stats['total_categorias']])
            ws_resumo.append(["Valor Total", f"R$ {stats['valor_total']:,.2f}"])
            ws_resumo.append(["Total de Unidades", stats['total_unidades']])
            ws_resumo.append([])
            
            # Resumo por categoria
            ws_resumo.append(["RESUMO POR CATEGORIAS"])
            ws_resumo.append(["Categoria", "Valor Total", "Quantidade", "Percentual"])
            
            for cat in analise['categorias']:
                ws_resumo.append([
                    cat['categoria'],
                    f"R$ {cat['total']:,.2f}",
                    cat['quantidade'],
                    f"{cat['percentual']:.1f}%"
                ])
            
            # === ABA CATEGORIAS DETALHADA ===
            ws_categorias = wb.create_sheet("Por Categorias")
            ws_categorias.append(["ANÁLISE DETALHADA POR CATEGORIAS"])
            ws_categorias.append([])
            
            for cat in analise['categorias']:
                ws_categorias.append([f"CATEGORIA: {cat['categoria']}"])
                ws_categorias.append([f"Total: R$ {cat['total']:,.2f} ({cat['percentual']:.1f}%)"])
                ws_categorias.append(["Procedimento", "Valor", "Quantidade"])
                
                for proc in cat['procedimentos']:
                    ws_categorias.append([
                        proc['procedimento'],
                        f"R$ {proc['total']:,.2f}",
                        proc['quantidade']
                    ])
                
                ws_categorias.append([])  # Linha em branco
            
            # === ABA PROCEDIMENTOS ===
            ws_procedimentos = wb.create_sheet("Por Procedimentos")
            ws_procedimentos.append(["ANÁLISE POR PROCEDIMENTOS"])
            ws_procedimentos.append([])
            ws_procedimentos.append(["Procedimento", "Categoria", "Valor Total", "Quantidade"])
            
            for proc in analise['procedimentos']:
                ws_procedimentos.append([
                    proc['procedimento'],
                    proc['categoria'],
                    f"R$ {proc['total']:,.2f}",
                    proc['quantidade']
                ])
            
            # === ABA UNIDADES ===
            ws_unidades = wb.create_sheet("Por Unidades")
            ws_unidades.append(["ANÁLISE POR UNIDADES"])
            ws_unidades.append([])
            ws_unidades.append(["Unidade", "Valor Total", "Quantidade", "Percentual"])
            
            for unidade in analise['unidades']:
                ws_unidades.append([
                    unidade['unidade'],
                    f"R$ {unidade['total']:,.2f}",
                    unidade['quantidade'],
                    f"{unidade['percentual']:.1f}%"
                ])
            
            # === ABA DADOS BRUTOS ===
            ws_dados = wb.create_sheet("Dados Brutos")
            ws_dados.append(["DADOS BRUTOS"])
            ws_dados.append([])
            ws_dados.append(["Unidade", "Procedimento", "Categoria", "Valor"])
            
            for _, row in df.iterrows():
                ws_dados.append([
                    row['Unidade'],
                    row['Procedimento'],
                    row['Categoria'],
                    f"R$ {row['TotalItem']:,.2f}"
                ])
            
            # Ajustar larguras das colunas
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 80)  # Máximo 80 para procedimentos longos
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Salvar
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return base64.b64encode(buffer.getvalue()).decode()
            
        except Exception as e:
            print(f"Erro ao gerar Excel: {e}")
            return None