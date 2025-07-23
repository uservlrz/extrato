from http.server import BaseHTTPRequestHandler
import json
import pandas as pd
import io
import base64
import openpyxl
import traceback

class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        
        response = {'status': 'OK', 'message': 'API de Procedimentos funcionando!'}
        self.wfile.write(json.dumps(response).encode())
    
    def do_POST(self):
        try:
            print("=== INICIANDO PROCESSAMENTO PROCEDIMENTOS ===")
            
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length)
            print(f"Dados recebidos: {len(post_data)} bytes")
            
            content_type = self.headers.get('Content-Type', '')
            if 'boundary=' not in content_type:
                raise Exception("Content-Type inválido - boundary não encontrado")
            
            boundary = content_type.split('boundary=')[1]
            print(f"Boundary: {boundary}")
            
            files, form_data = self.parse_multipart(post_data, boundary)
            print(f"Arquivos encontrados: {list(files.keys())}")
            print(f"Dados do formulário: {list(form_data.keys())}")
            
            procedures_data = files.get('procedures_file')
            categories_data = files.get('categories_file')
            
            if not procedures_data or not categories_data:
                raise Exception("Arquivos necessários não foram enviados")
            
            print(f"Procedures: {len(procedures_data)} bytes")
            print(f"Categories: {len(categories_data)} bytes")
            print("Processando todos os procedimentos")
            
            # Processar Categorias
            print("Processando categorias...")
            categorias = self.processar_arquivo_categorias(categories_data)
            print(f"Categorias encontradas: {len(categorias)}")
            
            # Processar Procedimentos
            print("Processando procedimentos...")
            df = self.processar_arquivo_procedimentos(procedures_data)
            print(f"Linhas processadas: {len(df)}")
            print(f"Colunas: {list(df.columns)}")
            
            # Categorizar
            print("Categorizando procedimentos...")
            df['Categoria'] = df['Procedimento'].apply(lambda x: self.mapear_procedimento_para_categoria(x, categorias))
            
            print(f"Total procedimentos: {len(df)} registros")
            
            # Agrupar resultados GERAIS
            print("Agrupando resultados gerais...")
            resultados_gerais = df.groupby('Categoria').agg({
                'TotalItem': ['sum', 'count']
            }).reset_index()
            resultados_gerais.columns = ['categoria', 'total', 'quantidade']
            valor_total = df['TotalItem'].sum()
            
            if valor_total > 0:
                resultados_gerais['percentual'] = (resultados_gerais['total'] / valor_total) * 100
            else:
                resultados_gerais['percentual'] = 0
            resultados_gerais = resultados_gerais.sort_values('total', ascending=False)
            
            # Agrupar resultados por PROCEDIMENTO
            print("Agrupando resultados por procedimento...")
            resultados_procedimentos = df.groupby('Procedimento').agg({
                'TotalItem': ['sum', 'count'],
                'Categoria': 'first'
            }).reset_index()
            resultados_procedimentos.columns = ['procedimento', 'total', 'quantidade', 'categoria']
            resultados_procedimentos = resultados_procedimentos.sort_values('total', ascending=False)
            
            # Agrupar resultados por UNIDADE
            print("Agrupando resultados por unidade...")
            resultados_unidades = df.groupby('Unidade').agg({
                'TotalItem': ['sum', 'count']
            }).reset_index()
            resultados_unidades.columns = ['unidade', 'total', 'quantidade']
            if valor_total > 0:
                resultados_unidades['percentual'] = (resultados_unidades['total'] / valor_total) * 100
            else:
                resultados_unidades['percentual'] = 0
            resultados_unidades = resultados_unidades.sort_values('total', ascending=False)
            
            # Função para preparar categorias detalhadas
            def preparar_categorias_detalhadas(resultados, dataframe, tipo):
                categorias_detalhadas = []
                for _, row in resultados.iterrows():
                    if tipo == 'categoria':
                        categoria = row['categoria']
                        itens_cat = dataframe[dataframe['Categoria'] == categoria]
                        
                        # Agrupar procedimentos desta categoria
                        procs_categoria = itens_cat.groupby('Procedimento').agg({
                            'TotalItem': ['sum', 'count']
                        }).reset_index()
                        procs_categoria.columns = ['procedimento', 'total', 'quantidade']
                        procs_categoria = procs_categoria.sort_values('total', ascending=False)
                        
                        procedimentos_lista = []
                        for _, proc in procs_categoria.iterrows():
                            procedimentos_lista.append({
                                'procedimento': proc['procedimento'],
                                'total': float(proc['total']),
                                'quantidade': int(proc['quantidade'])
                            })
                        
                        categorias_detalhadas.append({
                            'categoria': categoria,
                            'total': float(row['total']),
                            'quantidade': int(row['quantidade']),
                            'percentual': float(row['percentual']),
                            'procedimentos': procedimentos_lista
                        })
                        
                    elif tipo == 'procedimento':
                        procedimento = row['procedimento']
                        itens_proc = dataframe[dataframe['Procedimento'] == procedimento]
                        
                        # Agrupar unidades deste procedimento
                        unidades_proc = itens_proc.groupby('Unidade').agg({
                            'TotalItem': ['sum', 'count']
                        }).reset_index()
                        unidades_proc.columns = ['unidade', 'total', 'quantidade']
                        
                        unidades_dict = {}
                        for _, unidade in unidades_proc.iterrows():
                            unidades_dict[unidade['unidade']] = {
                                'total': float(unidade['total']),
                                'quantidade': int(unidade['quantidade'])
                            }
                        
                        categorias_detalhadas.append({
                            'procedimento': procedimento,
                            'categoria': row['categoria'],
                            'total': float(row['total']),
                            'quantidade': int(row['quantidade']),
                            'unidades': unidades_dict
                        })
                        
                    elif tipo == 'unidade':
                        unidade = row['unidade']
                        itens_unidade = dataframe[dataframe['Unidade'] == unidade]
                        
                        # Agrupar categorias desta unidade
                        cats_unidade = itens_unidade.groupby('Categoria').agg({
                            'TotalItem': ['sum', 'count']
                        }).reset_index()
                        cats_unidade.columns = ['categoria', 'total', 'quantidade']
                        cats_unidade = cats_unidade.sort_values('total', ascending=False)
                        
                        categorias_lista = []
                        for _, cat in cats_unidade.iterrows():
                            categorias_lista.append({
                                'categoria': cat['categoria'],
                                'total': float(cat['total']),
                                'quantidade': int(cat['quantidade'])
                            })
                        
                        categorias_detalhadas.append({
                            'unidade': unidade,
                            'total': float(row['total']),
                            'quantidade': int(row['quantidade']),
                            'percentual': float(row['percentual']),
                            'categorias': categorias_lista
                        })
                
                return categorias_detalhadas
            
            # Preparar respostas
            print("Preparando respostas...")
            categorias_gerais = preparar_categorias_detalhadas(resultados_gerais, df, 'categoria')
            procedimentos_detalhados = preparar_categorias_detalhadas(resultados_procedimentos, df, 'procedimento')
            unidades_detalhadas = preparar_categorias_detalhadas(resultados_unidades, df, 'unidade')
            
            # Gerar Excel
            print("Gerando Excel...")
            excel_b64 = self.gerar_excel_procedimentos(categorias_gerais, procedimentos_detalhados, unidades_detalhadas, df)
            
            resposta = {
                'success': True,
                'estatisticas': {
                    'total_procedimentos': len(df),
                    'total_categorias': len(resultados_gerais),
                    'valor_total': float(valor_total),
                    'total_unidades': len(resultados_unidades)
                },
                'categorias': categorias_gerais,
                'procedimentos': procedimentos_detalhados,
                'unidades': unidades_detalhadas,
                'excel_file': excel_b64
            }
            
            print("Enviando resposta...")
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(resposta).encode())
            print("=== PROCESSAMENTO CONCLUÍDO ===")
            
        except Exception as e:
            print(f"ERRO: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            error_response = {
                'success': False, 
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            self.wfile.write(json.dumps(error_response).encode())

    def parse_multipart(self, body, boundary):
        parts = body.split(f'--{boundary}'.encode())
        files = {}
        form_data = {}
        
        for part in parts:
            if b'Content-Disposition' not in part:
                continue
            
            header_end = part.find(b'\r\n\r\n')
            if header_end == -1:
                continue
            
            header = part[:header_end].decode('utf-8', errors='ignore')
            content = part[header_end + 4:].rstrip(b'\r\n-')
            
            if 'name="' in header:
                name_start = header.find('name="') + 6
                name_end = header.find('"', name_start)
                name = header[name_start:name_end]
                
                if 'filename="' in header:
                    files[name] = content
                else:
                    form_data[name] = content.decode('utf-8', errors='ignore')
        
        return files, form_data

    def processar_arquivo_categorias(self, categories_data):
        """Processa arquivo Excel de categorias"""
        try:
            df = pd.read_excel(io.BytesIO(categories_data))
            
            # Extrair lista de categorias da primeira coluna
            categorias = []
            for _, row in df.iterrows():
                if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip():
                    categorias.append(str(row.iloc[0]).strip())
            
            print(f"Categorias carregadas: {categorias}")
            return categorias
            
        except Exception as e:
            raise Exception(f"Erro ao processar arquivo de categorias: {e}")

    def processar_arquivo_procedimentos(self, procedures_data):
        """Processa arquivo Excel de procedimentos - Estrutura IPG"""
        try:
            print("Processando arquivo de procedimentos IPG...")
            
            # Ler o arquivo Excel
            df_raw = pd.read_excel(io.BytesIO(procedures_data))
            print(f"Arquivo carregado: {df_raw.shape}")
            
            # Verificar estrutura esperada do IPG
            # Linha 0: vazia, Linha 1: título, Linha 2: cabeçalhos
            print("Primeiras 5 linhas do arquivo:")
            for i in range(min(5, len(df_raw))):
                primeira_coluna = str(df_raw.iloc[i, 0]) if pd.notna(df_raw.iloc[i, 0]) else "VAZIO"
                print(f"  Linha {i}: {primeira_coluna}")
            
            # Tentar encontrar os cabeçalhos
            header_row = None
            for i in range(min(5, len(df_raw))):
                # Verificar se essa linha contém "Unidade" na primeira coluna
                first_cell = str(df_raw.iloc[i, 0]).strip().lower() if pd.notna(df_raw.iloc[i, 0]) else ""
                if 'unidade' in first_cell:
                    header_row = i
                    print(f"Cabeçalhos encontrados na linha {i}")
                    break
            
            if header_row is None:
                # Fallback: assumir que os cabeçalhos estão na linha 2 (padrão IPG)
                header_row = 2
                print("Usando linha 2 como cabeçalho (padrão IPG)")
            
            # Recarregar com o cabeçalho correto
            df = pd.read_excel(io.BytesIO(procedures_data), header=header_row)
            print(f"Dados recarregados com header na linha {header_row}: {df.shape}")
            print(f"Colunas encontradas: {list(df.columns)}")
            
            # Mapear as colunas específicas do IPG
            # Estrutura conhecida: Unidade (col 0), Procedimento (col 5), Total do Item (col 10)
            colunas_mapeadas = {}
            
            for i, col_name in enumerate(df.columns):
                col_str = str(col_name).strip()
                print(f"  Coluna {i}: '{col_str}'")
                
                # Mapear Unidade
                if i == 0 or 'unidade' in col_str.lower():
                    colunas_mapeadas['Unidade'] = col_name
                
                # Mapear Procedimento  
                elif i == 5 or 'procedimento' in col_str.lower():
                    colunas_mapeadas['Procedimento'] = col_name
                
                # Mapear Total do Item
                elif i == 10 or 'total do item' in col_str.lower() or 'total item' in col_str.lower():
                    colunas_mapeadas['TotalItem'] = col_name
            
            print(f"Colunas mapeadas: {colunas_mapeadas}")
            
            # Verificar se encontrou as 3 colunas essenciais
            if len(colunas_mapeadas) < 3:
                print("Não conseguiu mapear todas as colunas pelos nomes, usando posições fixas...")
                
                # Usar posições fixas baseadas na estrutura conhecida do IPG
                if df.shape[1] >= 11:  # Precisa ter pelo menos 11 colunas
                    df_normalizado = df.iloc[:, [0, 5, 10]].copy()
                    df_normalizado.columns = ['Unidade', 'Procedimento', 'TotalItem']
                    print("Usando posições fixas: coluna 0, 5, 10")
                else:
                    raise Exception(f"Arquivo não tem a estrutura esperada. Esperadas 11+ colunas, encontradas: {df.shape[1]}")
            else:
                # Usar as colunas mapeadas
                cols_needed = ['Unidade', 'Procedimento', 'TotalItem']
                missing_cols = [col for col in cols_needed if col not in colunas_mapeadas]
                
                if missing_cols:
                    raise Exception(f"Colunas não encontradas: {missing_cols}. Colunas disponíveis: {list(df.columns)}")
                
                df_normalizado = df[[colunas_mapeadas[col] for col in cols_needed]].copy()
                df_normalizado.columns = cols_needed
            
            print(f"Dados antes da limpeza: {len(df_normalizado)} registros")
            
            # Mostrar amostra dos dados brutos
            print("Amostra dos dados brutos:")
            print(df_normalizado.head().to_string())
            
            # Limpeza dos dados
            # 1. Remover linhas com valores nulos nas colunas essenciais
            df_normalizado = df_normalizado.dropna(subset=['Procedimento', 'TotalItem'])
            print(f"Após remover nulos: {len(df_normalizado)} registros")
            
            # 2. Remover linhas onde Procedimento está vazio
            df_normalizado = df_normalizado[df_normalizado['Procedimento'].astype(str).str.strip() != '']
            print(f"Após remover procedimentos vazios: {len(df_normalizado)} registros")
            
            # 3. Converter TotalItem para numérico
            def converter_valor(valor):
                if pd.isna(valor):
                    return 0.0
                try:
                    # Se já for número, retornar
                    if isinstance(valor, (int, float)):
                        return float(valor)
                    
                    # Se for string, limpar e converter
                    valor_str = str(valor).strip()
                    # Remover símbolos de moeda e espaços
                    valor_str = valor_str.replace('R$', '').replace('$', '').replace(' ', '')
                    # Trocar vírgula por ponto se necessário
                    if ',' in valor_str and '.' not in valor_str:
                        valor_str = valor_str.replace(',', '.')
                    
                    return float(valor_str)
                except:
                    return 0.0
            
            df_normalizado['TotalItem'] = df_normalizado['TotalItem'].apply(converter_valor)
            print(f"Valores convertidos - Min: {df_normalizado['TotalItem'].min()}, Max: {df_normalizado['TotalItem'].max()}")
            
            # 4. Remover registros com valor zero ou negativo
            df_normalizado = df_normalizado[df_normalizado['TotalItem'] > 0]
            print(f"Após filtrar valores > 0: {len(df_normalizado)} registros")
            
            # 5. Limpar campo Unidade
            df_normalizado['Unidade'] = df_normalizado['Unidade'].fillna('Não informado')
            df_normalizado['Unidade'] = df_normalizado['Unidade'].astype(str).str.strip()
            
            # 6. Limpar campo Procedimento
            df_normalizado['Procedimento'] = df_normalizado['Procedimento'].astype(str).str.strip()
            
            if len(df_normalizado) == 0:
                raise Exception("Nenhum registro válido encontrado após o processamento")
            
            print(f"✅ Processamento concluído: {len(df_normalizado)} registros válidos")
            
            # Mostrar estatísticas finais
            print(f"Unidades únicas: {df_normalizado['Unidade'].nunique()}")
            print(f"Procedimentos únicos: {df_normalizado['Procedimento'].nunique()}")
            print(f"Valor total: R$ {df_normalizado['TotalItem'].sum():,.2f}")
            
            # Mostrar amostra dos dados finais
            print("\nAmostra dos dados processados:")
            print(df_normalizado.head().to_string())
            
            return df_normalizado
            
        except Exception as e:
            print(f"❌ Erro detalhado no processamento: {e}")
            print(f"Traceback: {traceback.format_exc()}")
            raise Exception(f"Erro ao processar arquivo de procedimentos: {e}")

    def mapear_procedimento_para_categoria(self, procedimento, categorias):
        """Mapeia procedimento para categoria usando palavras-chave"""
        if not procedimento:
            return "Outros"
        
        proc_upper = str(procedimento).upper()
        
        # Buscar correspondência com categorias
        for categoria in categorias:
            cat_upper = str(categoria).upper().strip()
            
            # Verificar se a categoria está contida no procedimento
            if cat_upper in proc_upper:
                return categoria
            
            # Verificações específicas adicionais
            if cat_upper == 'VITAMINA' and ('VITAMINA' in proc_upper or 'B12' in proc_upper):
                return categoria
            if cat_upper == 'CONSULTAS' and ('CONSULTA' in proc_upper):
                return categoria
        
        return "Outros"

    def gerar_excel_procedimentos(self, categorias_gerais, procedimentos_detalhados, unidades_detalhadas, df):
        """Gera Excel com relatório de procedimentos - Mesmo padrão do extrato"""
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # === ABA RESUMO GERAL ===
            ws_resumo = wb.create_sheet("Resumo Geral")
            ws_resumo.append(["ANÁLISE COMPLETA DE PROCEDIMENTOS MÉDICOS"])
            ws_resumo.append([f"Gerado em: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"])
            ws_resumo.append([])
            
            # Estatísticas gerais
            total_procedimentos = len(df)
            total_categorias = len(categorias_gerais)
            total_unidades = len(unidades_detalhadas)
            valor_total = df['TotalItem'].sum()
            
            ws_resumo.append(["ESTATÍSTICAS GERAIS"])
            ws_resumo.append(["Total de Procedimentos", total_procedimentos])
            ws_resumo.append(["Total de Categorias", total_categorias])
            ws_resumo.append(["Total de Unidades", total_unidades])
            ws_resumo.append(["Valor Total Geral", f"R$ {valor_total:,.2f}"])
            ws_resumo.append([])
            
            # Resumo por categoria - GERAL
            ws_resumo.append(["RESUMO GERAL POR CATEGORIA"])
            ws_resumo.append(["Categoria", "Valor Total", "Quantidade", "Percentual"])
            
            for resultado in categorias_gerais:
                ws_resumo.append([
                    resultado['categoria'],
                    f"R$ {resultado['total']:,.2f}",
                    resultado['quantidade'],
                    f"{resultado['percentual']:.1f}%"
                ])
            
            # === ABA RESUMO PROCEDIMENTOS ===
            ws_procedimentos = wb.create_sheet("Resumo Procedimentos")
            ws_procedimentos.append(["ANÁLISE POR PROCEDIMENTOS"])
            ws_procedimentos.append([f"Total de Procedimentos: {len(procedimentos_detalhados)} tipos únicos"])
            ws_procedimentos.append([f"Valor Total: R$ {valor_total:,.2f}"])
            ws_procedimentos.append([])
            
            ws_procedimentos.append(["PROCEDIMENTOS POR CATEGORIA"])
            ws_procedimentos.append(["Procedimento", "Categoria", "Valor Total", "Quantidade"])
            
            for resultado in procedimentos_detalhados:
                ws_procedimentos.append([
                    resultado['procedimento'],
                    resultado['categoria'],
                    f"R$ {resultado['total']:,.2f}",
                    resultado['quantidade']
                ])
            
            # === ABA RESUMO UNIDADES ===
            ws_unidades = wb.create_sheet("Resumo Unidades")
            ws_unidades.append(["ANÁLISE POR UNIDADES"])
            ws_unidades.append([f"Total de Unidades: {len(unidades_detalhadas)} unidades"])
            ws_unidades.append([f"Valor Total: R$ {valor_total:,.2f}"])
            ws_unidades.append([])
            
            ws_unidades.append(["UNIDADES POR CATEGORIA"])
            ws_unidades.append(["Unidade", "Valor Total", "Quantidade", "Percentual"])
            
            for resultado in unidades_detalhadas:
                ws_unidades.append([
                    resultado['unidade'],
                    f"R$ {resultado['total']:,.2f}",
                    resultado['quantidade'],
                    f"{resultado['percentual']:.1f}%"
                ])
            
            # === ABAS DETALHADAS POR CATEGORIA ===
            
            # Função para criar aba detalhada
            def criar_aba_categoria(resultado, prefixo=""):
                categoria = resultado['categoria']
                
                # Nome da aba (máximo 31 caracteres, sem caracteres especiais)
                nome_aba = f"{prefixo}{categoria}".replace('/', '-').replace('\\', '-').replace('*', '-')
                nome_aba = nome_aba.replace('?', '').replace(':', '-').replace('[', '').replace(']', '')
                nome_aba = nome_aba[:31]  # Limite do Excel
                
                ws_categoria = wb.create_sheet(nome_aba)
                
                # Cabeçalho da categoria
                ws_categoria.append([f"CATEGORIA: {categoria}"])
                ws_categoria.append([f"Total: R$ {resultado['total']:,.2f}"])
                ws_categoria.append([f"Quantidade: {resultado['quantidade']} procedimentos"])
                ws_categoria.append([f"Percentual: {resultado['percentual']:.1f}% do total"])
                ws_categoria.append([])
                
                # Cabeçalho da tabela de procedimentos
                ws_categoria.append(["#", "Procedimento", "Valor", "Quantidade"])
                
                # Procedimentos da categoria
                for i, proc in enumerate(resultado['procedimentos'], 1):
                    ws_categoria.append([
                        i,
                        proc['procedimento'],
                        f"R$ {proc['total']:,.2f}",
                        proc['quantidade']
                    ])
                
                # Total da categoria
                ws_categoria.append([])
                ws_categoria.append(["", "TOTAL DA CATEGORIA:", f"R$ {resultado['total']:,.2f}", ""])
                
                # Ajustar largura das colunas
                ws_categoria.column_dimensions['A'].width = 5
                ws_categoria.column_dimensions['B'].width = 60
                ws_categoria.column_dimensions['C'].width = 15
                ws_categoria.column_dimensions['D'].width = 12
            
            # Criar abas para categorias
            for resultado in categorias_gerais:
                criar_aba_categoria(resultado)
            
            # Ajustar largura das colunas dos resumos
            for ws in [ws_resumo]:
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 12
            
            ws_procedimentos.column_dimensions['A'].width = 60
            ws_procedimentos.column_dimensions['B'].width = 25
            ws_procedimentos.column_dimensions['C'].width = 15
            ws_procedimentos.column_dimensions['D'].width = 12
            
            ws_unidades.column_dimensions['A'].width = 25
            ws_unidades.column_dimensions['B'].width = 15
            ws_unidades.column_dimensions['C'].width = 12
            ws_unidades.column_dimensions['D'].width = 12
            
            # Salvar
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return base64.b64encode(excel_buffer.getvalue()).decode()
            
        except Exception as e:
            print(f"Erro ao gerar Excel de procedimentos: {e}")
            return None