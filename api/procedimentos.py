from http.server import BaseHTTPRequestHandler
import json
import pandas as pd
import io
import base64
import openpyxl
import traceback

class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        
        response = {'status': 'OK', 'message': 'API de Procedimentos funcionando!'}
        self.wfile.write(json.dumps(response).encode())
    
    def do_POST(self):
        try:
            print("=== INICIANDO PROCESSAMENTO PROCEDIMENTOS ===")
            
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length)
            print(f"Dados recebidos: {len(post_data)} bytes")
            
            content_type = self.headers.get('Content-Type', '')
            if 'boundary=' not in content_type:
                raise Exception("Content-Type inválido - boundary não encontrado")
            
            boundary = content_type.split('boundary=')[1]
            print(f"Boundary: {boundary}")
            
            files, form_data = self.parse_multipart(post_data, boundary)
            print(f"Arquivos encontrados: {list(files.keys())}")
            print(f"Dados do formulário: {list(form_data.keys())}")
            
            procedures_data = files.get('procedures_file')
            categories_data = files.get('categories_file')
            
            if not procedures_data or not categories_data:
                raise Exception("Arquivos necessários não foram enviados")
            
            print(f"Procedures: {len(procedures_data)} bytes")
            print(f"Categories: {len(categories_data)} bytes")
            print("Processando todos os procedimentos")
            
            # Processar Categorias
            print("Processando categorias...")
            categorias = self.processar_arquivo_categorias(categories_data)
            print(f"Categorias encontradas: {len(categorias)}")
            
            # Processar Procedimentos
            print("Processando procedimentos...")
            df = self.processar_arquivo_procedimentos(procedures_data)
            print(f"Linhas processadas: {len(df)}")
            print(f"Colunas: {list(df.columns)}")
            
            # Categorizar
            print("Categorizando procedimentos...")
            df['Categoria'] = df['Procedimento'].apply(lambda x: self.mapear_procedimento_para_categoria(x, categorias))
            
            print(f"Total procedimentos: {len(df)} registros")
            
            # Agrupar resultados GERAIS
            print("Agrupando resultados gerais...")
            resultados_gerais = df.groupby('Categoria').agg({
                'TotalItem': ['sum', 'count']
            }).reset_index()
            resultados_gerais.columns = ['categoria', 'total', 'quantidade']
            valor_total = df['TotalItem'].sum()
            
            if valor_total > 0:
                resultados_gerais['percentual'] = (resultados_gerais['total'] / valor_total) * 100
            else:
                resultados_gerais['percentual'] = 0
            resultados_gerais = resultados_gerais.sort_values('total', ascending=False)
            
            # Agrupar resultados por PROCEDIMENTO
            print("Agrupando resultados por procedimento...")
            resultados_procedimentos = df.groupby('Procedimento').agg({
                'TotalItem': ['sum', 'count'],
                'Categoria': 'first'
            }).reset_index()
            resultados_procedimentos.columns = ['procedimento', 'total', 'quantidade', 'categoria']
            resultados_procedimentos = resultados_procedimentos.sort_values('total', ascending=False)
            
            # Agrupar resultados por UNIDADE
            print("Agrupando resultados por unidade...")
            resultados_unidades = df.groupby('Unidade').agg({
                'TotalItem': ['sum', 'count']
            }).reset_index()
            resultados_unidades.columns = ['unidade', 'total', 'quantidade']
            if valor_total > 0:
                resultados_unidades['percentual'] = (resultados_unidades['total'] / valor_total) * 100
            else:
                resultados_unidades['percentual'] = 0
            resultados_unidades = resultados_unidades.sort_values('total', ascending=False)
            
            # Função para preparar categorias detalhadas
            def preparar_categorias_detalhadas(resultados, dataframe, tipo):
                categorias_detalhadas = []
                for _, row in resultados.iterrows():
                    if tipo == 'categoria':
                        categoria = row['categoria']
                        itens_cat = dataframe[dataframe['Categoria'] == categoria]
                        
                        # Agrupar procedimentos desta categoria
                        procs_categoria = itens_cat.groupby('Procedimento').agg({
                            'TotalItem': ['sum', 'count']
                        }).reset_index()
                        procs_categoria.columns = ['procedimento', 'total', 'quantidade']
                        procs_categoria = procs_categoria.sort_values('total', ascending=False)
                        
                        procedimentos_lista = []
                        for _, proc in procs_categoria.iterrows():
                            procedimentos_lista.append({
                                'procedimento': proc['procedimento'],
                                'total': float(proc['total']),
                                'quantidade': int(proc['quantidade'])
                            })
                        
                        categorias_detalhadas.append({
                            'categoria': categoria,
                            'total': float(row['total']),
                            'quantidade': int(row['quantidade']),
                            'percentual': float(row['percentual']),
                            'procedimentos': procedimentos_lista
                        })
                        
                    elif tipo == 'procedimento':
                        procedimento = row['procedimento']
                        itens_proc = dataframe[dataframe['Procedimento'] == procedimento]
                        
                        # Agrupar unidades deste procedimento
                        unidades_proc = itens_proc.groupby('Unidade').agg({
                            'TotalItem': ['sum', 'count']
                        }).reset_index()
                        unidades_proc.columns = ['unidade', 'total', 'quantidade']
                        
                        unidades_dict = {}
                        for _, unidade in unidades_proc.iterrows():
                            unidades_dict[unidade['unidade']] = {
                                'total': float(unidade['total']),
                                'quantidade': int(unidade['quantidade'])
                            }
                        
                        categorias_detalhadas.append({
                            'procedimento': procedimento,
                            'categoria': row['categoria'],
                            'total': float(row['total']),
                            'quantidade': int(row['quantidade']),
                            'unidades': unidades_dict
                        })
                        
                    elif tipo == 'unidade':
                        unidade = row['unidade']
                        itens_unidade = dataframe[dataframe['Unidade'] == unidade]
                        
                        # Agrupar categorias desta unidade
                        cats_unidade = itens_unidade.groupby('Categoria').agg({
                            'TotalItem': ['sum', 'count']
                        }).reset_index()
                        cats_unidade.columns = ['categoria', 'total', 'quantidade']
                        cats_unidade = cats_unidade.sort_values('total', ascending=False)
                        
                        categorias_lista = []
                        for _, cat in cats_unidade.iterrows():
                            categorias_lista.append({
                                'categoria': cat['categoria'],
                                'total': float(cat['total']),
                                'quantidade': int(cat['quantidade'])
                            })
                        
                        categorias_detalhadas.append({
                            'unidade': unidade,
                            'total': float(row['total']),
                            'quantidade': int(row['quantidade']),
                            'percentual': float(row['percentual']),
                            'categorias': categorias_lista
                        })
                
                return categorias_detalhadas
            
            # Preparar respostas
            print("Preparando respostas...")
            categorias_gerais = preparar_categorias_detalhadas(resultados_gerais, df, 'categoria')
            procedimentos_detalhados = preparar_categorias_detalhadas(resultados_procedimentos, df, 'procedimento')
            unidades_detalhadas = preparar_categorias_detalhadas(resultados_unidades, df, 'unidade')
            
            # Gerar Excel
            print("Gerando Excel...")
            excel_b64 = self.gerar_excel_procedimentos(categorias_gerais, procedimentos_detalhados, unidades_detalhadas, df)
            
            resposta = {
                'success': True,
                'estatisticas': {
                    'total_procedimentos': len(df),
                    'total_categorias': len(resultados_gerais),
                    'valor_total': float(valor_total),
                    'total_unidades': len(resultados_unidades)
                },
                'categorias': categorias_gerais,
                'procedimentos': procedimentos_detalhados,
                'unidades': unidades_detalhadas,
                'excel_file': excel_b64
            }
            
            print("Enviando resposta...")
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(resposta).encode())
            print("=== PROCESSAMENTO CONCLUÍDO ===")
            
        except Exception as e:
            print(f"ERRO: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            error_response = {
                'success': False, 
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            self.wfile.write(json.dumps(error_response).encode())

    def parse_multipart(self, body, boundary):
        parts = body.split(f'--{boundary}'.encode())
        files = {}
        form_data = {}
        
        for part in parts:
            if b'Content-Disposition' not in part:
                continue
            
            header_end = part.find(b'\r\n\r\n')
            if header_end == -1:
                continue
            
            header = part[:header_end].decode('utf-8', errors='ignore')
            content = part[header_end + 4:].rstrip(b'\r\n-')
            
            if 'name="' in header:
                name_start = header.find('name="') + 6
                name_end = header.find('"', name_start)
                name = header[name_start:name_end]
                
                if 'filename="' in header:
                    files[name] = content
                else:
                    form_data[name] = content.decode('utf-8', errors='ignore')
        
        return files, form_data

    def processar_arquivo_categorias(self, categories_data):
        """Processa arquivo Excel de categorias de forma mais robusta"""
        try:
            df = pd.read_excel(io.BytesIO(categories_data))
            print(f"Arquivo de categorias carregado: {df.shape}")
            
            categorias = []
            
            # Tentar diferentes abordagens para extrair categorias
            for col_idx in range(min(3, df.shape[1])):  # Testar primeiras 3 colunas
                for _, row in df.iterrows():
                    if pd.notna(row.iloc[col_idx]):
                        valor = str(row.iloc[col_idx]).strip()
                        if valor and len(valor) > 1 and valor not in categorias:
                            # Filtrar valores que parecem ser categorias válidas
                            if not valor.isdigit() and not valor.startswith('Unnamed'):
                                categorias.append(valor)
            
            # Se encontrou poucas categorias, tentar uma abordagem mais agressiva
            if len(categorias) < 5:
                print("Poucas categorias encontradas, expandindo busca...")
                for col in df.columns:
                    for _, row in df.iterrows():
                        if pd.notna(row[col]):
                            valor = str(row[col]).strip()
                            if (valor and len(valor) > 2 and valor not in categorias 
                                and not valor.isdigit() and not valor.startswith('Unnamed')):
                                categorias.append(valor)
            
            # Remover duplicatas e limpar
            categorias = list(set(categorias))
            categorias = [cat for cat in categorias if len(cat) > 2]
            
            print(f"Categorias finais encontradas ({len(categorias)}): {categorias[:10]}...")
            return categorias
            
        except Exception as e:
            print(f"Erro ao processar categorias: {e}")
            # Retornar categorias padrão como fallback
            return ["CONSULTAS", "EXAMES", "PROCEDIMENTOS", "MEDICAMENTOS", "OUTROS"]

    def processar_arquivo_procedimentos(self, procedures_data):
        """Versão mais robusta para processar arquivo de procedimentos"""
        try:
            print("=== PROCESSAMENTO ROBUSTO DE PROCEDIMENTOS ===")
            
            # Primeiro, vamos ler o arquivo sem assumir nada sobre cabeçalhos
            df_raw = pd.read_excel(io.BytesIO(procedures_data), header=None)
            print(f"Arquivo carregado sem cabeçalho: {df_raw.shape}")
            
            # Examinar todas as linhas para encontrar os dados
            print("=== ANÁLISE COMPLETA DO ARQUIVO ===")
            for i in range(min(15, len(df_raw))):
                linha_info = []
                for j in range(min(12, df_raw.shape[1])):
                    valor = df_raw.iloc[i, j]
                    if pd.notna(valor):
                        valor_str = str(valor).strip()[:30]  # Mostrar só primeiros 30 chars
                        linha_info.append(f"Col{j}: {valor_str}")
                    else:
                        linha_info.append(f"Col{j}: VAZIO")
                print(f"Linha {i}: {' | '.join(linha_info[:5])}...")  # Mostrar só primeiras 5 colunas
            
            # ESTRATÉGIA 1: Buscar palavras-chave para identificar colunas
            unidade_col = None
            procedimento_col = None
            valor_col = None
            
            print("\n=== BUSCANDO COLUNAS POR PALAVRAS-CHAVE ===")
            for i in range(min(10, len(df_raw))):
                for j in range(min(15, df_raw.shape[1])):
                    cell_value = str(df_raw.iloc[i, j]).upper().strip() if pd.notna(df_raw.iloc[i, j]) else ""
                    
                    # Buscar coluna Unidade
                    if ('UNIDADE' in cell_value or 'UNIT' in cell_value) and unidade_col is None:
                        unidade_col = j
                        print(f"✅ Coluna UNIDADE encontrada: Col {j} (Linha {i})")
                    
                    # Buscar coluna Procedimento
                    if ('PROCEDIMENTO' in cell_value or 'PROC' in cell_value or 'DESCRI' in cell_value) and procedimento_col is None:
                        procedimento_col = j
                        print(f"✅ Coluna PROCEDIMENTO encontrada: Col {j} (Linha {i})")
                    
                    # Buscar coluna Valor/Total
                    if ('TOTAL' in cell_value or 'VALOR' in cell_value or 'VLR' in cell_value) and 'ITEM' in cell_value and valor_col is None:
                        valor_col = j
                        print(f"✅ Coluna VALOR encontrada: Col {j} (Linha {i})")
            
            # ESTRATÉGIA 2: Se não encontrou pelas palavras-chave, usar análise de dados
            if not all([unidade_col is not None, procedimento_col is not None, valor_col is not None]):
                print("\n=== ESTRATÉGIA ALTERNATIVA: ANÁLISE DE DADOS ===")
                
                # Analisar cada coluna para determinar seu tipo
                for j in range(min(15, df_raw.shape[1])):
                    sample_data = []
                    for i in range(5, min(25, len(df_raw))):  # Pular primeiras linhas (cabeçalhos)
                        if pd.notna(df_raw.iloc[i, j]):
                            sample_data.append(str(df_raw.iloc[i, j]).strip())
                    
                    if len(sample_data) < 3:
                        continue
                    
                    # Verificar se é coluna de valor (números)
                    numeric_count = 0
                    for val in sample_data[:10]:
                        try:
                            # Tentar converter para float
                            val_clean = val.replace(',', '.').replace('R$', '').replace(' ', '')
                            float(val_clean)
                            numeric_count += 1
                        except:
                            pass
                    
                    if numeric_count > len(sample_data) * 0.7 and valor_col is None:
                        valor_col = j
                        print(f"📊 Coluna VALOR identificada por análise: Col {j}")
                    
                    # Verificar se é coluna de texto longo (procedimentos)
                    if numeric_count < len(sample_data) * 0.3:
                        avg_length = sum(len(val) for val in sample_data) / len(sample_data)
                        if avg_length > 15 and procedimento_col is None:  # Textos longos = procedimentos
                            procedimento_col = j
                            print(f"📝 Coluna PROCEDIMENTO identificada por análise: Col {j}")
                        elif avg_length < 15 and unidade_col is None:  # Textos curtos = unidades
                            unidade_col = j
                            print(f"🏢 Coluna UNIDADE identificada por análise: Col {j}")
            
            # ESTRATÉGIA 3: Fallback para posições conhecidas
            if not all([unidade_col is not None, procedimento_col is not None, valor_col is not None]):
                print("\n=== USANDO POSIÇÕES PADRÃO COMO FALLBACK ===")
                if unidade_col is None:
                    unidade_col = 0
                    print(f"🔄 Usando Col 0 para UNIDADE")
                if procedimento_col is None:
                    procedimento_col = 5 if df_raw.shape[1] > 5 else min(2, df_raw.shape[1]-1)
                    print(f"🔄 Usando Col {procedimento_col} para PROCEDIMENTO")
                if valor_col is None:
                    valor_col = min(10, df_raw.shape[1]-1)
                    print(f"🔄 Usando Col {valor_col} para VALOR")
            
            print(f"\n✅ COLUNAS FINAIS: Unidade=Col{unidade_col}, Procedimento=Col{procedimento_col}, Valor=Col{valor_col}")
            
            # Extrair os dados usando as colunas identificadas
            dados_extraidos = []
            linhas_processadas = 0
            
            print("\n=== EXTRAINDO DADOS ===")
            for i in range(len(df_raw)):
                try:
                    unidade = df_raw.iloc[i, unidade_col] if unidade_col < df_raw.shape[1] else None
                    procedimento = df_raw.iloc[i, procedimento_col] if procedimento_col < df_raw.shape[1] else None
                    valor = df_raw.iloc[i, valor_col] if valor_col < df_raw.shape[1] else None
                    
                    # Pular linhas de cabeçalho e vazias
                    if (pd.isna(procedimento) or pd.isna(valor) or 
                        str(procedimento).upper().strip() in ['PROCEDIMENTO', 'DESCRICAO', 'PROC', ''] or
                        str(valor).upper().strip() in ['TOTAL', 'VALOR', 'VLR', '']):
                        continue
                    
                    # Limpar e validar dados
                    unidade_clean = str(unidade).strip() if pd.notna(unidade) else "Não informado"
                    procedimento_clean = str(procedimento).strip()
                    
                    # Converter valor
                    valor_clean = self.converter_valor_robusto(valor)
                    
                    if len(procedimento_clean) > 3 and valor_clean > 0:
                        dados_extraidos.append({
                            'Unidade': unidade_clean,
                            'Procedimento': procedimento_clean,
                            'TotalItem': valor_clean
                        })
                        linhas_processadas += 1
                        
                        # Log das primeiras 5 linhas para debug
                        if linhas_processadas <= 5:
                            print(f"  Linha {i}: {unidade_clean[:20]}... | {procedimento_clean[:30]}... | R$ {valor_clean:,.2f}")
                
                except Exception as e:
                    continue
            
            if len(dados_extraidos) == 0:
                raise Exception("Nenhum dado válido encontrado no arquivo")
            
            # Criar DataFrame final
            df_final = pd.DataFrame(dados_extraidos)
            
            print(f"\n✅ PROCESSAMENTO CONCLUÍDO:")
            print(f"   📊 Total de registros: {len(df_final)}")
            print(f"   🏢 Unidades únicas: {df_final['Unidade'].nunique()}")
            print(f"   📝 Procedimentos únicos: {df_final['Procedimento'].nunique()}")
            print(f"   💰 Valor total: R$ {df_final['TotalItem'].sum():,.2f}")
            
            # Mostrar amostra dos dados
            print(f"\n📋 AMOSTRA DOS DADOS PROCESSADOS:")
            print(df_final.head().to_string())
            
            return df_final
            
        except Exception as e:
            print(f"❌ ERRO CRÍTICO no processamento: {e}")
            print(f"Traceback completo: {traceback.format_exc()}")
            raise Exception(f"Erro ao processar arquivo de procedimentos: {e}")

    def converter_valor_robusto(self, valor):
        """Converte valores de forma mais robusta"""
        if pd.isna(valor):
            return 0.0
        
        try:
            # Se já for número
            if isinstance(valor, (int, float)):
                return float(valor)
            
            # Se for string, limpar completamente
            valor_str = str(valor).strip()
            
            # Remover todos os caracteres não numéricos exceto , e .
            import re
            valor_clean = re.sub(r'[^\d,.-]', '', valor_str)
            
            # Tratar vírgulas e pontos
            if ',' in valor_clean and '.' in valor_clean:
                # Formato: 1.234,56 -> 1234.56
                if valor_clean.rfind(',') > valor_clean.rfind('.'):
                    valor_clean = valor_clean.replace('.', '').replace(',', '.')
                else:
                    # Formato: 1,234.56 -> 1234.56
                    valor_clean = valor_clean.replace(',', '')
            elif ',' in valor_clean:
                # Se só tem vírgula, assumir que é decimal
                valor_clean = valor_clean.replace(',', '.')
            
            resultado = float(valor_clean)
            return abs(resultado)  # Sempre positivo
            
        except Exception as e:
            print(f"⚠️ Erro ao converter valor '{valor}': {e}")
            return 0.0

    def mapear_procedimento_para_categoria(self, procedimento, categorias):
        """Mapeia procedimento para categoria de forma mais inteligente"""
        if not procedimento:
            return "Outros"
        
        proc_upper = str(procedimento).upper()
        
        # Mapeamento específico primeiro
        mapeamentos_especificos = {
            'CONSULTA': 'CONSULTAS',
            'EXAM': 'EXAMES',
            'ULTRA': 'EXAMES',
            'RAIO': 'EXAMES',
            'TOMOGRAFIA': 'EXAMES',
            'RESSONANCIA': 'EXAMES',
            'SANGUE': 'EXAMES',
            'VITAMINA': 'MEDICAMENTOS',
            'MEDICAMENTO': 'MEDICAMENTOS',
            'DIPIRONA': 'MEDICAMENTOS',
            'CIRURGIA': 'PROCEDIMENTOS',
            'PROCEDIMENTO': 'PROCEDIMENTOS'
        }
        
        # Verificar mapeamentos específicos primeiro
        for palavra_chave, categoria_destino in mapeamentos_especificos.items():
            if palavra_chave in proc_upper:
                return categoria_destino
        
        # Depois verificar as categorias fornecidas
        for categoria in categorias:
            cat_upper = str(categoria).upper().strip()
            if cat_upper and cat_upper in proc_upper:
                return categoria
        
        return "Outros"

    def gerar_excel_procedimentos(self, categorias_gerais, procedimentos_detalhados, unidades_detalhadas, df):
        """Gera Excel com relatório de procedimentos"""
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # === ABA RESUMO GERAL ===
            ws_resumo = wb.create_sheet("Resumo Geral")
            ws_resumo.append(["ANÁLISE COMPLETA DE PROCEDIMENTOS MÉDICOS"])
            ws_resumo.append([f"Gerado em: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"])
            ws_resumo.append([])
            
            # Estatísticas gerais
            total_procedimentos = len(df)
            total_categorias = len(categorias_gerais)
            total_unidades = len(unidades_detalhadas)
            valor_total = df['TotalItem'].sum()
            
            ws_resumo.append(["ESTATÍSTICAS GERAIS"])
            ws_resumo.append(["Total de Procedimentos", total_procedimentos])
            ws_resumo.append(["Total de Categorias", total_categorias])
            ws_resumo.append(["Total de Unidades", total_unidades])
            ws_resumo.append(["Valor Total Geral", f"R$ {valor_total:,.2f}"])
            ws_resumo.append([])
            
            # Resumo por categoria - GERAL
            ws_resumo.append(["RESUMO GERAL POR CATEGORIA"])
            ws_resumo.append(["Categoria", "Valor Total", "Quantidade", "Percentual"])
            
            for resultado in categorias_gerais:
                ws_resumo.append([
                    resultado['categoria'],
                    f"R$ {resultado['total']:,.2f}",
                    resultado['quantidade'],
                    f"{resultado['percentual']:.1f}%"
                ])
            
            # === ABA RESUMO PROCEDIMENTOS ===
            ws_procedimentos = wb.create_sheet("Resumo Procedimentos")
            ws_procedimentos.append(["ANÁLISE POR PROCEDIMENTOS"])
            ws_procedimentos.append([f"Total de Procedimentos: {len(procedimentos_detalhados)} tipos únicos"])
            ws_procedimentos.append([f"Valor Total: R$ {valor_total:,.2f}"])
            ws_procedimentos.append([])
            
            ws_procedimentos.append(["PROCEDIMENTOS POR CATEGORIA"])
            ws_procedimentos.append(["Procedimento", "Categoria", "Valor Total", "Quantidade"])
            
            for resultado in procedimentos_detalhados:
                ws_procedimentos.append([
                    resultado['procedimento'],
                    resultado['categoria'],
                    f"R$ {resultado['total']:,.2f}",
                    resultado['quantidade']
                ])
            
            # === ABA RESUMO UNIDADES ===
            ws_unidades = wb.create_sheet("Resumo Unidades")
            ws_unidades.append(["ANÁLISE POR UNIDADES"])
            ws_unidades.append([f"Total de Unidades: {len(unidades_detalhadas)} unidades"])
            ws_unidades.append([f"Valor Total: R$ {valor_total:,.2f}"])
            ws_unidades.append([])
            
            ws_unidades.append(["UNIDADES POR CATEGORIA"])
            ws_unidades.append(["Unidade", "Valor Total", "Quantidade", "Percentual"])
            
            for resultado in unidades_detalhadas:
                ws_unidades.append([
                    resultado['unidade'],
                    f"R$ {resultado['total']:,.2f}",
                    resultado['quantidade'],
                    f"{resultado['percentual']:.1f}%"
                ])
            
            # === ABAS DETALHADAS POR CATEGORIA ===
            def criar_aba_categoria(resultado, prefixo=""):
                categoria = resultado['categoria']
                
                nome_aba = f"{prefixo}{categoria}".replace('/', '-').replace('\\', '-').replace('*', '-')
                nome_aba = nome_aba.replace('?', '').replace(':', '-').replace('[', '').replace(']', '')
                nome_aba = nome_aba[:31]  # Limite do Excel
                
                ws_categoria = wb.create_sheet(nome_aba)
                
                # Cabeçalho da categoria
                ws_categoria.append([f"CATEGORIA: {categoria}"])
                ws_categoria.append([f"Total: R$ {resultado['total']:,.2f}"])
                ws_categoria.append([f"Quantidade: {resultado['quantidade']} procedimentos"])
                ws_categoria.append([f"Percentual: {resultado['percentual']:.1f}% do total"])
                ws_categoria.append([])
                
                # Cabeçalho da tabela de procedimentos
                ws_categoria.append(["#", "Procedimento", "Valor", "Quantidade"])
                
                # Procedimentos da categoria
                for i, proc in enumerate(resultado['procedimentos'], 1):
                    ws_categoria.append([
                        i,
                        proc['procedimento'],
                        f"R$ {proc['total']:,.2f}",
                        proc['quantidade']
                    ])
                
                # Total da categoria
                ws_categoria.append([])
                ws_categoria.append(["", "TOTAL DA CATEGORIA:", f"R$ {resultado['total']:,.2f}", ""])
                
                # Ajustar largura das colunas
                ws_categoria.column_dimensions['A'].width = 5
                ws_categoria.column_dimensions['B'].width = 60
                ws_categoria.column_dimensions['C'].width = 15
                ws_categoria.column_dimensions['D'].width = 12
            
            # Criar abas para categorias
            for resultado in categorias_gerais:
                criar_aba_categoria(resultado)
            
            # === ABA DADOS BRUTOS ===
            ws_dados = wb.create_sheet("Dados Brutos")
            ws_dados.append(["TODOS OS DADOS PROCESSADOS"])
            ws_dados.append([])
            ws_dados.append(["#", "Unidade", "Procedimento", "Categoria", "Valor"])
            
            for i, row in df.iterrows():
                ws_dados.append([
                    i + 1,
                    row['Unidade'],
                    row['Procedimento'],
                    row['Categoria'],
                    f"R$ {row['TotalItem']:,.2f}"
                ])
            
            # Ajustar largura das colunas
            ws_dados.column_dimensions['A'].width = 5
            ws_dados.column_dimensions['B'].width = 25
            ws_dados.column_dimensions['C'].width = 60
            ws_dados.column_dimensions['D'].width = 20
            ws_dados.column_dimensions['E'].width = 15
            
            # Ajustar largura das colunas dos resumos
            for ws in [ws_resumo]:
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 12
            
            ws_procedimentos.column_dimensions['A'].width = 60
            ws_procedimentos.column_dimensions['B'].width = 25
            ws_procedimentos.column_dimensions['C'].width = 15
            ws_procedimentos.column_dimensions['D'].width = 12
            
            ws_unidades.column_dimensions['A'].width = 25
            ws_unidades.column_dimensions['B'].width = 15
            ws_unidades.column_dimensions['C'].width = 12
            ws_unidades.column_dimensions['D'].width = 12
            
            # Salvar
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            print(f"✅ Excel gerado com {len(wb.worksheets)} abas")
            return base64.b64encode(excel_buffer.getvalue()).decode()
            
        except Exception as e:
            print(f"❌ Erro ao gerar Excel: {e}")
            print(f"Traceback: {traceback.format_exc()}")
            return None