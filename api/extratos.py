from http.server import BaseHTTPRequestHandler
import json
import pandas as pd
import io
import base64
import openpyxl
import re
import traceback

class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        
        response = {'status': 'OK', 'message': 'API funcionando!'}
        self.wfile.write(json.dumps(response).encode())
    
    def do_POST(self):
        try:
            print("=== INICIANDO PROCESSAMENTO ===")
            
            # Receber dados
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length)
            
            # Parse multipart
            content_type = self.headers.get('Content-Type', '')
            if 'boundary=' not in content_type:
                raise Exception("Content-Type inválido - boundary não encontrado")
            
            boundary = content_type.split('boundary=')[1]
            files, form_data = self.parse_multipart(post_data, boundary)
            
            csv_data = files.get('csv_file')
            excel_data = files.get('excel_file')
            
            if not csv_data or not excel_data:
                raise Exception("Arquivos necessários não foram enviados")
            
            print(f"CSV: {len(csv_data)} bytes, Excel: {len(excel_data)} bytes")
            
            # Processar arquivos
            categorias = self.processar_excel(excel_data)
            df = self.processar_csv(csv_data)
            
            # Categorizar transações
            df['Categoria'] = df['Descricao'].apply(lambda x: self.categorizar(x, categorias))
            
            # Separar por tipo
            df_creditos = df[df['Tipo'] == 'C'].copy()
            df_debitos = df[df['Tipo'] == 'D'].copy()
            
            # Gerar resultados
            resultados = self.gerar_resultados(df, df_creditos, df_debitos)
            
            # Gerar Excel
            excel_b64 = self.gerar_excel_completo(
                resultados['categorias_gerais'], 
                resultados['categorias_creditos'], 
                resultados['categorias_debitos'], 
                df, df_creditos, df_debitos
            )
            
            # Resposta final
            resposta = {
                'success': True,
                'estatisticas': resultados['estatisticas'],
                'categorias_gerais': resultados['categorias_gerais'],
                'categorias_creditos': resultados['categorias_creditos'],
                'categorias_debitos': resultados['categorias_debitos'],
                'excel_file': excel_b64
            }
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(resposta).encode())
            
            print("=== PROCESSAMENTO CONCLUÍDO ===")
            
        except Exception as e:
            print(f"ERRO: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            error_response = {
                'success': False, 
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            self.wfile.write(json.dumps(error_response).encode())

    # ==========================================
    # UTILITÁRIOS
    # ==========================================
    
    def parse_multipart(self, body, boundary):
        """Parse de dados multipart/form-data"""
        parts = body.split(f'--{boundary}'.encode())
        files = {}
        form_data = {}
        
        for part in parts:
            if b'Content-Disposition' not in part:
                continue
            
            header_end = part.find(b'\r\n\r\n')
            if header_end == -1:
                continue
            
            header = part[:header_end].decode('utf-8', errors='ignore')
            content = part[header_end + 4:].rstrip(b'\r\n-')
            
            if 'name="' in header:
                name_start = header.find('name="') + 6
                name_end = header.find('"', name_start)
                name = header[name_start:name_end]
                
                if 'filename="' in header:
                    files[name] = content
                else:
                    form_data[name] = content.decode('utf-8', errors='ignore')
        
        return files, form_data

    def processar_valor_monetario(self, valor):
        """Converte valores monetários brasileiros para float"""
        if pd.isna(valor) or valor == '' or valor is None:
            return 0.0
        
        # Converter para string e limpar espaços
        valor_str = str(valor).strip()
        if not valor_str or valor_str.lower() == 'nan':
            return 0.0
        
        # Remover aspas se houver
        valor_str = valor_str.replace('"', '').replace("'", "")
        
        # Se está vazio após limpeza
        if not valor_str:
            return 0.0
        
        # Tratar valores negativos (podem vir com hífen ou entre aspas negativas)
        negativo = False
        if valor_str.startswith('-') or valor_str.startswith('"-'):
            negativo = True
            valor_str = valor_str.lstrip('-').lstrip('"').rstrip('"')
        
        # Formato brasileiro: remover pontos (milhares) e trocar vírgula por ponto
        # Exemplo: "1.234,56" -> "1234.56"
        if ',' in valor_str and '.' in valor_str:
            # Se tem ambos, ponto é milhares e vírgula é decimal
            valor_str = valor_str.replace('.', '').replace(',', '.')
        elif ',' in valor_str:
            # Só vírgula, é decimal
            valor_str = valor_str.replace(',', '.')
        # Se só tem ponto, assumir que é decimal (formato americano ou sem milhares)
        
        try:
            resultado = float(valor_str)
            # Aplicar sinal negativo se necessário
            return -resultado if negativo else resultado
        except ValueError as e:
            print(f"Erro ao processar valor '{valor}' -> '{valor_str}': {e}")
            return 0.0

    # ==========================================
    # PROCESSAMENTO EXCEL
    # ==========================================
    
    def processar_excel(self, excel_data):
        """Processa arquivo Excel de categorias"""
        try:
            print("=== PROCESSANDO EXCEL ===")
            
            # Verificar formato do arquivo
            formato = self.verificar_formato_excel(excel_data)
            print(f"Formato detectado: {formato}")
            
            if formato == 'xls':
                raise Exception("Arquivos .xls (Excel antigo) não são suportados. Por favor, abra o arquivo no Excel e salve como .xlsx")
            
            # Tentar processar como .xlsx
            try:
                df = pd.read_excel(io.BytesIO(excel_data), engine='openpyxl')
                print("Excel processado com sucesso (.xlsx)")
            except Exception as e1:
                print(f"Erro ao processar Excel: {e1}")
                # Tentar sem especificar engine
                try:
                    df = pd.read_excel(io.BytesIO(excel_data))
                    print("Excel processado com engine padrão")
                except Exception as e2:
                    raise Exception(f"Não foi possível processar o arquivo Excel. Certifique-se de que é um arquivo .xlsx válido. Erro: {e1}")
            
            print(f"Excel carregado: {len(df)} linhas, {len(df.columns)} colunas")
            
            if len(df.columns) < 2:
                raise Exception("Excel deve ter pelo menos 2 colunas (Grupo e Palavra-chave)")
            
            # Normalizar nomes das colunas
            df.columns = ['Grupo', 'Palavra_Chave'] + list(df.columns[2:])
            
            print("Estrutura do Excel:")
            print(f"  Colunas: {list(df.columns)}")
            
            # Processar categorias
            categorias = {}
            categoria_atual = None
            
            for index, row in df.iterrows():
                # Se tem grupo definido, usar como categoria atual
                if pd.notna(row['Grupo']) and str(row['Grupo']).strip():
                    categoria_atual = str(row['Grupo']).strip()
                
                # Se tem palavra-chave e categoria atual, adicionar
                if pd.notna(row['Palavra_Chave']) and categoria_atual:
                    palavra = str(row['Palavra_Chave']).strip()
                    if palavra:  # Só adicionar se não estiver vazio
                        categorias[palavra] = categoria_atual
            
            print(f"Total de categorias processadas: {len(categorias)}")
            
            if len(categorias) == 0:
                raise Exception("Nenhuma categoria válida encontrada no Excel. Verifique o formato do arquivo.")
            
            return categorias
            
        except Exception as e:
            print(f"Erro detalhado no Excel: {e}")
            raise Exception(f"Erro no Excel: {e}")
    
    def verificar_formato_excel(self, excel_data):
        """Verifica se é .xls ou .xlsx baseado nos primeiros bytes"""
        # .xlsx começa com PK (ZIP signature)
        if excel_data[:2] == b'PK':
            return 'xlsx'
        # .xls tem assinatura específica
        elif excel_data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
            return 'xls'
        else:
            return 'unknown'

    # ==========================================
    # PROCESSAMENTO CSV - VERSÃO SIMPLIFICADA
    # ==========================================
    
    def processar_csv(self, csv_data):
        """Processa CSV - versão universal simplificada"""
        try:
            print("=== PROCESSANDO CSV ===")
            
            # Decodificar CSV
            csv_string = None
            for encoding in ['utf-8', 'latin1', 'cp1252']:
                try:
                    csv_string = csv_data.decode(encoding)
                    print(f"CSV decodificado com {encoding}")
                    break
                except:
                    continue
            
            if not csv_string:
                raise Exception("Não foi possível decodificar o CSV")
            
            print(f"Tamanho do arquivo: {len(csv_string)} caracteres")
            
            # Detectar tipo básico (BB vs Bradesco)
            if self.eh_banco_brasil(csv_string):
                print("Formato detectado: Banco do Brasil")
                return self.processar_banco_brasil(csv_string)
            else:
                print("Formato detectado: Bradesco")
                return self.processar_bradesco_universal(csv_string)
                
        except Exception as e:
            print(f"Erro no processamento CSV: {e}")
            raise e
    
    def eh_banco_brasil(self, csv_string):
        """Verifica se é Banco do Brasil"""
        indicadores_bb = [
            '"DATA","DEPENDENCIA ORIGEM"',
            '"DATA","HISTÓRICO"',
            '"DATA","HISTORICO"'
        ]
        
        for indicador in indicadores_bb:
            if indicador in csv_string.upper():
                return True
        
        # Se tem muito mais vírgulas que ponto-vírgulas, provavelmente é BB
        virgulas = csv_string.count(',')
        ponto_virgulas = csv_string.count(';')
        
        return virgulas > ponto_virgulas * 2

    def processar_banco_brasil(self, csv_string):
        """Processa CSV do Banco do Brasil"""
        print("=== PROCESSANDO BANCO DO BRASIL ===")
        
        # Limpar caracteres problemáticos
        csv_string = csv_string.replace('Histórico', 'Historico').replace('Número', 'Numero')
        
        df = pd.read_csv(io.StringIO(csv_string))
        
        # Detectar formato e processar
        if 'Descrição' in df.columns or 'Descricao' in df.columns:
            desc_col = 'Descrição' if 'Descrição' in df.columns else 'Descricao'
            df = df.dropna(subset=[desc_col, 'Valor'])
            df['Descricao'] = df[desc_col]
            df['Documento'] = df.get('Documento', '')
        elif 'Historico' in df.columns:
            df = df.dropna(subset=['Historico', 'Valor'])
            df = df[df['Historico'] != 'Saldo Anterior']
            df['Descricao'] = df['Historico']
            df['Documento'] = df.get('Numero do documento', '')
            df['Tipo'] = df['Valor'].apply(lambda x: 'C' if x >= 0 else 'D')
            df['Valor'] = df['Valor'].abs()
        else:
            raise Exception("Formato de CSV do Banco do Brasil não reconhecido")
        
        df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')
        df = df.dropna(subset=['Valor'])
        
        print(f"Banco do Brasil processado: {len(df)} linhas")
        return df

    def processar_bradesco_universal(self, csv_string):
        """Processa qualquer formato de Bradesco de forma universal - VERSÃO ROBUSTA"""
        print("=== PROCESSANDO BRADESCO UNIVERSAL (VERSÃO ROBUSTA) ===")
        
        linhas = csv_string.split('\n')
        print(f"Total de linhas: {len(linhas)}")
        
        # Debug: mostrar estrutura
        print("=== ESTRUTURA DO ARQUIVO (primeiras 20 linhas) ===")
        for i, linha in enumerate(linhas[:20]):
            print(f"{i:2d}: {linha}")
        
        # TENTATIVA 1: Formato novo organizado (linhas individuais)
        print("\n=== TENTATIVA 1: FORMATO NOVO ORGANIZADO ===")
        try:
            resultado = self.processar_formato_novo_organizado(csv_string)
            if len(resultado) > 0:
                print(f"✅ SUCESSO - Formato novo: {len(resultado)} transações")
                return resultado
            else:
                print("❌ Formato novo não retornou dados")
        except Exception as e:
            print(f"❌ Erro formato novo: {e}")
        
        # TENTATIVA 2: Formato antigo (dados concatenados com \r)
        print("\n=== TENTATIVA 2: FORMATO ANTIGO (DADOS CONCATENADOS) ===")
        try:
            resultado = self.processar_formato_antigo_concatenado(csv_string)
            if len(resultado) > 0:
                print(f"✅ SUCESSO - Formato antigo: {len(resultado)} transações")
                return resultado
            else:
                print("❌ Formato antigo não retornou dados")
        except Exception as e:
            print(f"❌ Erro formato antigo: {e}")
        
        # TENTATIVA 3: Processamento linha por linha universal
        print("\n=== TENTATIVA 3: PROCESSAMENTO UNIVERSAL LINHA POR LINHA ===")
        try:
            transacoes = []
            
            for i, linha in enumerate(linhas):
                linha_limpa = linha.strip()
                if not linha_limpa:
                    continue
                
                if self.eh_transacao_bradesco(linha_limpa):
                    transacao = self.extrair_transacao_bradesco(linha_limpa)
                    if transacao:
                        transacoes.append(transacao)
                        if len(transacoes) <= 5:
                            print(f"Transação {len(transacoes)}: {transacao['Data']} - {transacao['Descricao'][:30]}... - R$ {transacao['Valor']}")
            
            if transacoes:
                df = pd.DataFrame(transacoes)
                df = self.processar_datas_padrao(df)
                print(f"✅ SUCESSO - Processamento universal: {len(df)} transações")
                return df
            else:
                print("❌ Processamento universal não encontrou transações")
                
        except Exception as e:
            print(f"❌ Erro processamento universal: {e}")
        
        # Se chegou até aqui, fazer debug detalhado
        print("\n=== NENHUM MÉTODO FUNCIONOU - DEBUG DETALHADO ===")
        self.debug_arquivo_bradesco(csv_string)
        raise Exception("Não foi possível processar o arquivo Bradesco com nenhum método")
    
    def processar_formato_novo_organizado(self, csv_string):
        """Processa formato novo com linhas organizadas"""
        linhas = csv_string.split('\n')
        
        # Procurar cabeçalho
        header_line = -1
        cabecalho = None
        
        padroes_cabecalho = [
            r'Data;.*?Histórico.*?;.*?Docto',
            r'Data;.*?Historico.*?;.*?Docto', 
            r'Data;.*?Histórico.*?;.*?Crédito',
            r'Data;.*?Historico.*?;.*?Credito'
        ]
        
        for i, linha in enumerate(linhas):
            for padrao in padroes_cabecalho:
                if re.search(padrao, linha, re.IGNORECASE):
                    header_line = i
                    cabecalho = linha.strip()
                    print(f"Cabeçalho encontrado linha {i}: {cabecalho}")
                    break
            if header_line != -1:
                break
        
        if header_line == -1:
            raise Exception("Cabeçalho formato novo não encontrado")
        
        # Extrair transações após cabeçalho
        transacoes_linhas = []
        for i in range(header_line + 1, len(linhas)):
            linha = linhas[i].strip()
            if not linha:
                continue
            
            # Parar em indicadores de fim
            if any(fim in linha.upper() for fim in ['OS DADOS ACIMA', 'TOTAL GERAL', 'GERADO EM']):
                break
            
            # Pular controles
            if any(ctrl in linha.upper() for ctrl in ['SALDO ANTERIOR', 'EXTRATO DE', 'ÚLTIMOS LANÇAMENTOS']):
                continue
            
            # Se começa com data e tem campos suficientes
            if re.match(r'^\d{2}/\d{2}/\d{2,4};', linha) and linha.count(';') >= 3:
                transacoes_linhas.append(linha)
        
        if not transacoes_linhas:
            return pd.DataFrame(columns=['Data', 'Descricao', 'Valor', 'Tipo', 'Documento'])
        
        # Criar CSV e processar
        csv_data = cabecalho + '\n' + '\n'.join(transacoes_linhas)
        
        try:
            df = pd.read_csv(io.StringIO(csv_data), delimiter=';')
        except Exception as e:
            print(f"Erro ao criar DataFrame: {e}")
            df = pd.read_csv(io.StringIO(csv_data), delimiter=';', on_bad_lines='skip')
        
        return self.processar_dataframe_bradesco(df)
    
    def processar_formato_antigo_concatenado(self, csv_string):
        """Processa formato antigo com dados concatenados em uma linha"""
        linhas = csv_string.split('\n')
        
        # Procurar linha com muitos \r (dados concatenados)
        linha_dados = None
        for linha in linhas:
            if linha.count('\r') > 10 and ';' in linha and 'Data;' in linha:
                linha_dados = linha
                break
        
        if not linha_dados:
            # Tentar concatenar linhas a partir do cabeçalho
            for i, linha in enumerate(linhas):
                if re.search(r'Data;.*?Lançamento.*?;.*?Dcto', linha, re.IGNORECASE):
                    linha_dados = ''.join(linhas[i:])
                    break
        
        if not linha_dados:
            raise Exception("Dados concatenados não encontrados")
        
        # Separar por \r
        partes = linha_dados.split('\r')
        cabecalho = partes[0].strip()
        
        # Filtrar transações válidas
        transacoes_linhas = []
        for parte in partes[1:]:
            linha = parte.strip()
            if (linha and 
                re.match(r'^\d{2}/\d{2}/\d{4};', linha) and
                linha.count(';') >= 4 and
                'SALDO ANTERIOR' not in linha.upper()):
                transacoes_linhas.append(linha)
        
        if not transacoes_linhas:
            return pd.DataFrame(columns=['Data', 'Descricao', 'Valor', 'Tipo', 'Documento'])
        
        # Criar CSV e processar
        csv_data = cabecalho + '\n' + '\n'.join(transacoes_linhas)
        
        try:
            df = pd.read_csv(io.StringIO(csv_data), delimiter=';')
        except Exception as e:
            print(f"Erro ao criar DataFrame: {e}")
            df = pd.read_csv(io.StringIO(csv_data), delimiter=';', on_bad_lines='skip')
        
        return self.processar_dataframe_bradesco(df)
    
    def processar_dataframe_bradesco(self, df):
        """Processa DataFrame do Bradesco de forma universal"""
        print(f"Processando DataFrame: {len(df)} linhas, colunas: {list(df.columns)}")
        
        # Mapear colunas de forma inteligente
        mapeamento = {}
        for col in df.columns:
            col_lower = col.lower().strip()
            if 'data' in col_lower:
                mapeamento[col] = 'Data'
            elif any(desc in col_lower for desc in ['histórico', 'historico', 'lançamento', 'lancamento']):
                mapeamento[col] = 'Descricao'
            elif any(doc in col_lower for doc in ['docto', 'documento']):
                mapeamento[col] = 'Documento'
            elif any(cred in col_lower for cred in ['crédito', 'credito']):
                mapeamento[col] = 'Credito'
            elif any(deb in col_lower for deb in ['débito', 'debito']):
                mapeamento[col] = 'Debito'
        
        df = df.rename(columns=mapeamento)
        print(f"Colunas após mapeamento: {list(df.columns)}")
        
        # Garantir colunas essenciais
        for col in ['Credito', 'Debito', 'Documento']:
            if col not in df.columns:
                df[col] = '' if col == 'Documento' else 0.0
        
        # Processar valores
        if 'Credito' in df.columns:
            df['Credito'] = df['Credito'].apply(self.processar_valor_monetario)
        if 'Debito' in df.columns:
            df['Debito'] = df['Debito'].apply(self.processar_valor_monetario)
        
        # Determinar tipo e valor
        def determinar_tipo_valor(row):
            credito = row.get('Credito', 0)
            debito = row.get('Debito', 0)
            
            if credito > 0:
                return 'C', credito
            elif debito != 0:
                return 'D', abs(debito)
            else:
                return 'D', 0.0
        
        df[['Tipo', 'Valor']] = df.apply(lambda row: pd.Series(determinar_tipo_valor(row)), axis=1)
        
        # Limpar dados
        df = df[df['Valor'] > 0].dropna(subset=['Descricao'])
        
        # Processar datas
        df = self.processar_datas_padrao(df)
        
        # Retornar colunas padrão
        colunas_resultado = ['Data', 'Descricao', 'Valor', 'Tipo', 'Documento']
        for col in colunas_resultado:
            if col not in df.columns:
                df[col] = '' if col == 'Documento' else None
        
        resultado = df[colunas_resultado].reset_index(drop=True)
        print(f"Resultado processado: {len(resultado)} linhas válidas")
        
        return resultado
    
    def processar_datas_padrao(self, df):
        """Processa datas de forma padrão"""
        if 'Data' not in df.columns:
            return df
        
        try:
            df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%y', errors='coerce')
            if df['Data'].isna().all():
                df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y', errors='coerce')
        except:
            print("Mantendo datas como string")
        
        return df

    def eh_transacao_bradesco(self, linha):
        """Verifica se uma linha é uma transação do Bradesco"""
        # Critérios para ser transação:
        criterios = 0
        
        # 1. Tem data brasileira
        if re.search(r'\d{2}/\d{2}/\d{2,4}', linha):
            criterios += 1
        
        # 2. Tem múltiplos campos
        if linha.count(';') >= 3:
            criterios += 1
        
        # 3. Não é linha de controle
        if not any(ctrl in linha.upper() for ctrl in [
            'SALDO ANTERIOR', 'EXTRATO DE', 'AGÊNCIA', 'CONTA', 'TOTAL',
            'OS DADOS ACIMA', 'DATA;HISTÓRICO', 'DATA;HISTORICO', 'DATA;LANÇAMENTO'
        ]):
            criterios += 1
        
        # 4. Tem conteúdo textual (descrição)
        if any(len(campo.strip()) > 5 for campo in linha.split(';') 
               if not re.match(r'^[\d\.,\-\s"]*$', campo.strip())):
            criterios += 1
        
        return criterios >= 3

    def extrair_transacao_bradesco(self, linha):
        """Extrai dados de uma transação do Bradesco"""
        campos = [campo.strip() for campo in linha.split(';')]
        
        # Extrair data
        data = None
        for campo in campos:
            match = re.search(r'\d{2}/\d{2}/\d{2,4}', campo)
            if match:
                data = match.group()
                break
        
        if not data:
            return None
        
        # Extrair descrição (maior campo textual)
        descricao = ""
        for campo in campos:
            campo_limpo = campo.replace('"', '').strip()
            if (len(campo_limpo) > len(descricao) and
                len(campo_limpo) > 3 and
                not re.match(r'^\d{2}/\d{2}/\d{2,4}$', campo_limpo) and
                not re.match(r'^[\d\.,\-\s]*$', campo_limpo)):
                descricao = campo_limpo
        
        if not descricao:
            return None
        
        # Extrair valores
        valores = []
        for campo in campos:
            valor = self.processar_valor_monetario(campo)
            if valor != 0:
                valores.append(valor)
        
        if not valores:
            return None
        
        # Determinar valor principal e tipo
        valor_principal = max(valores, key=abs)
        tipo = 'C' if valor_principal > 0 else 'D'
        valor_final = abs(valor_principal)
        
        return {
            'Data': data,
            'Descricao': descricao,
            'Valor': valor_final,
            'Tipo': tipo,
            'Documento': ''
        }

    def debug_arquivo_bradesco(self, csv_string):
        """Debug quando não consegue processar Bradesco"""
        print("\n=== DEBUG DETALHADO ===")
        
        linhas = csv_string.split('\n')
        
        print(f"Total de linhas: {len(linhas)}")
        print(f"Uso de ';': {csv_string.count(';')}")
        print(f"Uso de '\\r': {csv_string.count(chr(13))}")
        
        # Datas encontradas
        datas = re.findall(r'\d{2}/\d{2}/\d{2,4}', csv_string)
        print(f"Datas encontradas: {len(datas)} - {datas[:5] if datas else 'nenhuma'}")
        
        # Linhas com múltiplos campos
        linhas_complexas = [linha for linha in linhas if linha.count(';') >= 3]
        print(f"Linhas com 3+ campos: {len(linhas_complexas)}")
        
        for i, linha in enumerate(linhas_complexas[:5]):
            print(f"  {i+1}: {linha[:100]}...")

    # ==========================================
    # CATEGORIZAÇÃO E RESULTADOS
    # ==========================================
    
    def categorizar(self, descricao, categorias):
        """Categoriza uma descrição baseada nas palavras-chave"""
        if not descricao or pd.isna(descricao):
            return "Outros"
        
        desc_upper = str(descricao).upper()
        
        # Ordenar por tamanho decrescente para matches mais específicos primeiro
        sorted_keys = sorted(categorias.keys(), key=len, reverse=True)
        
        for keyword in sorted_keys:
            if keyword.upper() in desc_upper:
                return categorias[keyword]
        
        return "Outros"

    def gerar_resultados(self, df, df_creditos, df_debitos):
        """Gera todos os resultados agrupados"""
        def agrupar_por_categoria(dataframe, nome_tipo):
            if len(dataframe) == 0:
                return pd.DataFrame(columns=['categoria', 'total', 'quantidade', 'percentual'])
            
            resultados = dataframe.groupby('Categoria').agg({
                'Valor': ['sum', 'count']
            }).reset_index()
            resultados.columns = ['categoria', 'total', 'quantidade']
            
            valor_total = dataframe['Valor'].sum()
            if valor_total > 0:
                resultados['percentual'] = (resultados['total'] / valor_total) * 100
            else:
                resultados['percentual'] = 0
            
            resultados = resultados.sort_values('total', ascending=False)
            return resultados
        
        # Agrupar por categoria
        resultados_gerais = agrupar_por_categoria(df, "Geral")
        resultados_creditos = agrupar_por_categoria(df_creditos, "Créditos")
        resultados_debitos = agrupar_por_categoria(df_debitos, "Débitos")
        
        # Preparar categorias detalhadas
        def preparar_categorias_detalhadas(resultados, dataframe):
            categorias_detalhadas = []
            for _, row in resultados.iterrows():
                categoria = row['categoria']
                itens_cat = dataframe[dataframe['Categoria'] == categoria]
                
                itens = []
                for _, item in itens_cat.iterrows():
                    data_formatada = str(item['Data']) if not pd.isna(item['Data']) else None
                    
                    itens.append({
                        'data': data_formatada,
                        'descricao': str(item['Descricao']),
                        'valor': float(item['Valor']),
                        'tipo': str(item['Tipo']),
                        'documento': str(item.get('Documento', ''))
                    })
                
                categorias_detalhadas.append({
                    'categoria': categoria,
                    'total': float(row['total']),
                    'quantidade': int(row['quantidade']),
                    'percentual': float(row['percentual']),
                    'itens': itens
                })
            return categorias_detalhadas
        
        # Estatísticas
        estatisticas = {
            'total_transacoes': len(df),
            'total_debitos': len(df_debitos),
            'total_creditos': len(df_creditos),
            'valor_total': float(df['Valor'].sum()),
            'valor_total_creditos': float(df_creditos['Valor'].sum() if len(df_creditos) > 0 else 0),
            'valor_total_debitos': float(df_debitos['Valor'].sum() if len(df_debitos) > 0 else 0)
        }
        
        return {
            'estatisticas': estatisticas,
            'categorias_gerais': preparar_categorias_detalhadas(resultados_gerais, df),
            'categorias_creditos': preparar_categorias_detalhadas(resultados_creditos, df_creditos),
            'categorias_debitos': preparar_categorias_detalhadas(resultados_debitos, df_debitos)
        }

    # ==========================================
    # GERAÇÃO DE EXCEL
    # ==========================================
    
    def gerar_excel_completo(self, categorias_gerais, categorias_creditos, categorias_debitos, df_geral, df_creditos, df_debitos):
        """Gera Excel completo com todas as abas - VERSÃO COMPLETA RESTAURADA"""
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # Estatísticas
            total_transacoes = len(df_geral)
            total_debitos = len(df_debitos)
            total_creditos = len(df_creditos)
            valor_total = df_geral['Valor'].sum()
            valor_creditos = df_creditos['Valor'].sum() if len(df_creditos) > 0 else 0
            valor_debitos = df_debitos['Valor'].sum() if len(df_debitos) > 0 else 0
            
            # === ABA RESUMO GERAL ===
            ws_resumo = wb.create_sheet("Resumo Geral")
            ws_resumo.append(["ANÁLISE COMPLETA DE EXTRATO BANCÁRIO"])
            ws_resumo.append([f"Gerado em: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"])
            ws_resumo.append([])
            
            ws_resumo.append(["ESTATÍSTICAS GERAIS"])
            ws_resumo.append(["Total de Transações", total_transacoes])
            ws_resumo.append(["Total de Créditos", total_creditos])
            ws_resumo.append(["Total de Débitos", total_debitos])
            ws_resumo.append(["Valor Total Geral", f"R$ {valor_total:,.2f}"])
            ws_resumo.append(["Valor Total Créditos", f"R$ {valor_creditos:,.2f}"])
            ws_resumo.append(["Valor Total Débitos", f"R$ {valor_debitos:,.2f}"])
            ws_resumo.append(["Saldo (Créditos - Débitos)", f"R$ {(valor_creditos - valor_debitos):,.2f}"])
            ws_resumo.append([])
            
            # Resumo por categoria GERAL
            ws_resumo.append(["RESUMO GERAL POR CATEGORIA"])
            ws_resumo.append(["Categoria", "Valor Total", "Quantidade", "Percentual"])
            
            for resultado in categorias_gerais:
                ws_resumo.append([
                    resultado['categoria'],
                    f"R$ {resultado['total']:,.2f}",
                    resultado['quantidade'],
                    f"{resultado['percentual']:.1f}%"
                ])
            
            # === ABA RESUMO CRÉDITOS ===
            if len(categorias_creditos) > 0:
                ws_creditos = wb.create_sheet("Resumo Créditos")
                ws_creditos.append(["ANÁLISE DE CRÉDITOS (ENTRADAS)"])
                ws_creditos.append([f"Total de Créditos: {total_creditos} transações"])
                ws_creditos.append([f"Valor Total: R$ {valor_creditos:,.2f}"])
                ws_creditos.append([])
                
                ws_creditos.append(["CRÉDITOS POR CATEGORIA"])
                ws_creditos.append(["Categoria", "Valor Total", "Quantidade", "Percentual"])
                
                for resultado in categorias_creditos:
                    ws_creditos.append([
                        resultado['categoria'],
                        f"R$ {resultado['total']:,.2f}",
                        resultado['quantidade'],
                        f"{resultado['percentual']:.1f}%"
                    ])
            
            # === ABA RESUMO DÉBITOS ===
            if len(categorias_debitos) > 0:
                ws_debitos = wb.create_sheet("Resumo Débitos")
                ws_debitos.append(["ANÁLISE DE DÉBITOS (SAÍDAS)"])
                ws_debitos.append([f"Total de Débitos: {total_debitos} transações"])
                ws_debitos.append([f"Valor Total: R$ {valor_debitos:,.2f}"])
                ws_debitos.append([])
                
                ws_debitos.append(["DÉBITOS POR CATEGORIA"])
                ws_debitos.append(["Categoria", "Valor Total", "Quantidade", "Percentual"])
                
                for resultado in categorias_debitos:
                    ws_debitos.append([
                        resultado['categoria'],
                        f"R$ {resultado['total']:,.2f}",
                        resultado['quantidade'],
                        f"{resultado['percentual']:.1f}%"
                    ])
            
            # === FUNÇÃO PARA CRIAR ABAS DETALHADAS ===
            def criar_aba_categoria(resultado, prefixo=""):
                categoria = resultado['categoria']
                
                # Nome da aba (máximo 31 caracteres, sem caracteres especiais)
                nome_aba = f"{prefixo}{categoria}".replace('/', '-').replace('\\', '-').replace('*', '-')
                nome_aba = nome_aba.replace('?', '').replace(':', '-').replace('[', '').replace(']', '')
                nome_aba = nome_aba[:31]  # Limite do Excel
                
                ws_categoria = wb.create_sheet(nome_aba)
                
                # Cabeçalho da categoria
                ws_categoria.append([f"CATEGORIA: {categoria}"])
                ws_categoria.append([f"Total: R$ {resultado['total']:,.2f}"])
                ws_categoria.append([f"Quantidade: {resultado['quantidade']} itens"])
                ws_categoria.append([f"Percentual: {resultado['percentual']:.1f}% do total"])
                ws_categoria.append([])
                
                # Cabeçalho da tabela de itens
                ws_categoria.append(["#", "Data", "Descrição", "Valor", "Tipo", "Documento"])
                
                # Itens da categoria
                for i, item in enumerate(resultado['itens'], 1):
                    # Formatar data
                    if item['data']:
                        try:
                            if isinstance(item['data'], str):
                                data_formatada = pd.to_datetime(item['data'], dayfirst=True).strftime('%d/%m/%Y')
                            else:
                                data_formatada = item['data'].strftime('%d/%m/%Y')
                        except:
                            data_formatada = str(item['data'])
                    else:
                        data_formatada = 'Sem data'
                    
                    # Formatar tipo
                    tipo_formatado = "CRÉDITO" if item['tipo'] == 'C' else "DÉBITO"
                    
                    ws_categoria.append([
                        i,
                        data_formatada,
                        item['descricao'],
                        f"R$ {item['valor']:,.2f}",
                        tipo_formatado,
                        str(item['documento'])
                    ])
                
                # Total da categoria
                ws_categoria.append([])
                ws_categoria.append(["", "", "TOTAL DA CATEGORIA:", f"R$ {resultado['total']:,.2f}", "", ""])
                
                # Ajustar largura das colunas
                ws_categoria.column_dimensions['A'].width = 5
                ws_categoria.column_dimensions['B'].width = 12
                ws_categoria.column_dimensions['C'].width = 50
                ws_categoria.column_dimensions['D'].width = 15
                ws_categoria.column_dimensions['E'].width = 10
                ws_categoria.column_dimensions['F'].width = 15
            
            # Criar abas para categorias gerais (principais)
            for resultado in categorias_gerais:
                criar_aba_categoria(resultado)
            
            # Criar abas para créditos (se houver)
            for resultado in categorias_creditos:
                criar_aba_categoria(resultado, "C_")
            
            # Criar abas para débitos (se houver)
            for resultado in categorias_debitos:
                criar_aba_categoria(resultado, "D_")
            
            # Ajustar largura das colunas dos resumos
            ws_resumo.column_dimensions['A'].width = 25
            ws_resumo.column_dimensions['B'].width = 15
            ws_resumo.column_dimensions['C'].width = 12
            ws_resumo.column_dimensions['D'].width = 12
            
            if len(categorias_creditos) > 0:
                ws_creditos.column_dimensions['A'].width = 25
                ws_creditos.column_dimensions['B'].width = 15
                ws_creditos.column_dimensions['C'].width = 12
                ws_creditos.column_dimensions['D'].width = 12
            
            if len(categorias_debitos) > 0:
                ws_debitos.column_dimensions['A'].width = 25
                ws_debitos.column_dimensions['B'].width = 15
                ws_debitos.column_dimensions['C'].width = 12
                ws_debitos.column_dimensions['D'].width = 12
            
            # Salvar Excel
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return base64.b64encode(excel_buffer.getvalue()).decode()
            
        except Exception as e:
            print(f"Erro ao gerar Excel: {e}")
            return None